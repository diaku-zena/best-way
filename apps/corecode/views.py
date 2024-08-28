from datetime import datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.shortcuts import HttpResponseRedirect, redirect, render
from django.urls import reverse_lazy
from django.views.generic import ListView, TemplateView, View
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.models import User

from .forms import (
    AcademicSessionForm,
    AcademicTermForm,
    CurrentSessionForm,
    SiteConfigForm,
    StudentClassForm, 
    SubjectForm,
    PermitDocCategoryForm,
    CitizenshipForm,
    DocumentTypeForm,
    FuncaoChefiaForm,
    CategoriaForm,
    DepartamentoForm,
    ExcelUploadForm,
)
from .models import (
    AcademicSession,
    AcademicTerm,
    CategoriaNova,
    SiteConfig,
    StudentClass,
    Subject,
    PermitDocCategory,
    Citizenship,
    DocumentType,
    FuncaoChefia,
    Categoria,
    Departamento,
    DirecaoAlocacao
)

from apps.employees.models import Employee

from apps.prova_vida.models import ProvaVida
from apps.prova_vida.models import Abertura_Prova_Vida

def create_user(request):
    if not User.objects.filter(username='admin').exists():
        user = User.objects.create_superuser("admin", "admin2@bestway.com", "admin123")
    print("creATE UER")
    template_name = "registration/login.html"

    return render(request, template_name)



class IndexView(LoginRequiredMixin, TemplateView):
    template_name = "index.html"


class SiteConfigView(LoginRequiredMixin, View):
    """Site Config View"""

    form_class = SiteConfigForm
    template_name = "corecode/siteconfig.html"

    def get(self, request, *args, **kwargs):
        formset = self.form_class(queryset=SiteConfig.objects.all())
        context = {"formset": formset}
        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        formset = self.form_class(request.POST)
        if formset.is_valid():
            formset.save()
            messages.success(request, "Configurations successfully updated")
        context = {"formset": formset, "title": "Configuration"}
        return render(request, self.template_name, context)


class SessionListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = AcademicSession
    template_name = "corecode/session_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AcademicSessionForm()
        return context


class SessionCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = AcademicSession
    form_class = AcademicSessionForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("sessions")
    success_message = "New session successfully added"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["title"] = "Add new session"
        return context


class SessionUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = AcademicSession
    form_class = AcademicSessionForm
    success_url = reverse_lazy("sessions")
    success_message = "Session successfully updated."
    template_name = "corecode/mgt_form.html"

    def form_valid(self, form):
        obj = self.object
        if obj.current == False:
            terms = (
                AcademicSession.objects.filter(current=True)
                .exclude(name=obj.name)
                .exists()
            )
            if not terms:
                messages.warning(self.request, "You must set a session to current.")
                return redirect("session-list")
        return super().form_valid(form)


class SessionDeleteView(LoginRequiredMixin, DeleteView):
    model = AcademicSession
    success_url = reverse_lazy("sessions")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "The session {} has been deleted with all its attached content"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.current == True:
            messages.warning(request, "Cannot delete session as it is set to current")
            return redirect("sessions")
        messages.success(self.request, self.success_message.format(obj.name))
        return super(SessionDeleteView, self).delete(request, *args, **kwargs)

class TermListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = AcademicTerm
    template_name = "corecode/term_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AcademicTermForm()
        return context


class TermCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = AcademicTerm
    form_class = AcademicTermForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("terms")
    success_message = "New term successfully added"


class TermUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = AcademicTerm
    form_class = AcademicTermForm
    success_url = reverse_lazy("terms")
    success_message = "Term successfully updated."
    template_name = "corecode/mgt_form.html"

    def form_valid(self, form):
        obj = self.object
        if obj.current == False:
            terms = (
                AcademicTerm.objects.filter(current=True)
                .exclude(name=obj.name)
                .exists()
            )
            if not terms:
                messages.warning(self.request, "You must set a term to current.")
                return redirect("term")
        return super().form_valid(form)


class TermDeleteView(LoginRequiredMixin, DeleteView):
    model = AcademicTerm
    success_url = reverse_lazy("terms")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "The term {} has been deleted with all its attached content"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.current == True:
            messages.warning(request, "Cannot delete term as it is set to current")
            return redirect("terms")
        messages.success(self.request, self.success_message.format(obj.name))
        return super(TermDeleteView, self).delete(request, *args, **kwargs)


class ClassListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = StudentClass
    template_name = "corecode/class_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = StudentClassForm()
        return context


class ClassCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = StudentClass
    form_class = StudentClassForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("classes")
    success_message = "New class successfully added"


class ClassUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = StudentClass
    fields = ["name"]
    success_url = reverse_lazy("classes")
    success_message = "class successfully updated."
    template_name = "corecode/mgt_form.html"


class ClassDeleteView(LoginRequiredMixin, DeleteView):
    model = StudentClass
    success_url = reverse_lazy("classes")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "The class {} has been deleted with all its attached content"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        print(obj.name)
        messages.success(self.request, self.success_message.format(obj.name))
        return super(ClassDeleteView, self).delete(request, *args, **kwargs)

class PermitDocCategoryListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = PermitDocCategory
    template_name = "corecode/permit_doc_cat_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = PermitDocCategoryForm()
        return context


class PermitDocCategoryCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = PermitDocCategory
    form_class = PermitDocCategoryForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("permitdoccategory")
    success_message = "Новый документ успешно добавлен"


class PermitDocCategoryUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = PermitDocCategory
    fields = ["name"]
    success_url = reverse_lazy("permitdoccategory")
    success_message = "Документ успешно обновлен."
    template_name = "corecode/mgt_form.html"


class PermitDocCategoryDeleteView(LoginRequiredMixin, DeleteView):
    model = PermitDocCategory
    success_url = reverse_lazy("permitdoccategory")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "Документ {} был удален со всем прикрепленным к нему содержимым"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        print(obj.name)
        messages.success(self.request, self.success_message.format(obj.name))
        return super(PermitDocCategoryDeleteView, self).delete(request, *args, **kwargs)

class CitizenshipListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = Citizenship
    template_name = "corecode/citizenship_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = CitizenshipForm()
        return context


class CitizenshipCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Citizenship
    form_class = CitizenshipForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("citizenship")
    success_message = "Новое гражданство успешно добавлено"


class CitizenshipUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Citizenship
    fields = ["name"]
    success_url = reverse_lazy("citizenship")
    success_message = "Гражданство успешно обновлено."
    template_name = "corecode/mgt_form.html"


class CitizenshipDeleteView(LoginRequiredMixin, DeleteView):
    model = Citizenship
    success_url = reverse_lazy("citizenship")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "Гражданство {} был удален со всем прикрепленным к нему содержимым"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        print(obj.name)
        messages.success(self.request, self.success_message.format(obj.name))
        return super(CitizenshipDeleteView, self).delete(request, *args, **kwargs)


class DocumentTypeListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = DocumentType
    template_name = "corecode/doc_type_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = DocumentTypeForm()
        return context


class DocumentTypeCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = DocumentType
    form_class = DocumentTypeForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("doctype")
    success_message = "Новый тип документа успешно добавлено"


class DocumentTypeUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = DocumentType
    fields = ["name"]
    success_url = reverse_lazy("doctype")
    success_message = "Тип документа успешно обновлено."
    template_name = "corecode/mgt_form.html"


class DocumentTypeDeleteView(LoginRequiredMixin, DeleteView):
    model = DocumentType
    success_url = reverse_lazy("doctype")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "Тип документа {} был удален со всем прикрепленным к нему содержимым"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        print(obj.name)
        messages.success(self.request, self.success_message.format(obj.name))
        return super(DocumentTypeDeleteView, self).delete(request, *args, **kwargs)


class SubjectListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = Subject
    template_name = "corecode/subject_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = SubjectForm()
        return context


class SubjectCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Subject
    form_class = SubjectForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("subjects")
    success_message = "New subject successfully added"


class SubjectUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Subject
    fields = ["name"]
    success_url = reverse_lazy("subjects")
    success_message = "Subject successfully updated."
    template_name = "corecode/mgt_form.html"


class SubjectDeleteView(LoginRequiredMixin, DeleteView):
    model = Subject
    success_url = reverse_lazy("subjects")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "The subject {} has been deleted with all its attached content"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        messages.success(self.request, self.success_message.format(obj.name))
        return super(SubjectDeleteView, self).delete(request, *args, **kwargs)


class CurrentSessionAndTermView(LoginRequiredMixin, View):
    """Current SEssion and Term"""

    form_class = CurrentSessionForm
    template_name = "corecode/current_session.html"

    def get(self, request, *args, **kwargs):
        form = self.form_class(
            initial={
                "current_session": AcademicSession.objects.get(current=True),
                "current_term": AcademicTerm.objects.get(current=True),
            }
        )
        return render(request, self.template_name, {"form": form})

    def post(self, request, *args, **kwargs):
        form = self.form_Class(request.POST)
        if form.is_valid():
            session = form.cleaned_data["current_session"]
            term = form.cleaned_data["current_term"]
            AcademicSession.objects.filter(name=session).update(current=True)
            AcademicSession.objects.exclude(name=session).update(current=False)
            AcademicTerm.objects.filter(name=term).update(current=True)

        return render(request, self.template_name, {"form": form})





#catgoria,funcao,
    

class FuncaoChefiaListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = FuncaoChefia
    template_name = "corecode/funcao_chefia_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AcademicTermForm()
        return context


class FuncaoChefiaCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = FuncaoChefia
    form_class = AcademicTermForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("terms")
    success_message = "New term successfully added"


class FuncaoChefiaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = FuncaoChefia
    form_class = AcademicTermForm
    success_url = reverse_lazy("terms")
    success_message = "Term successfully updated."
    template_name = "corecode/mgt_form.html"

    def form_valid(self, form):
        obj = self.object
        if obj.current == False:
            terms = (
                AcademicTerm.objects.filter(current=True)
                .exclude(name=obj.name)
                .exists()
            )
            if not terms:
                messages.warning(self.request, "You must set a term to current.")
                return redirect("term")
        return super().form_valid(form)
    


class FuncaoChefiaListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = FuncaoChefia
    template_name = "corecode/funcao_chefia_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AcademicTermForm()
        return context


class FuncaoChefiaCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = FuncaoChefia
    form_class = AcademicTermForm
    template_name = "corecode/mgt_form.html"
    success_url = reverse_lazy("terms")
    success_message = "New term successfully added"


class FuncaoChefiaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = FuncaoChefia
    form_class = AcademicTermForm
    success_url = reverse_lazy("terms")
    success_message = "Term successfully updated."
    template_name = "corecode/mgt_form.html"

    def form_valid(self, form):
        obj = self.object
        if obj.current == False:
            terms = (
                AcademicTerm.objects.filter(current=True)
                .exclude(name=obj.name)
                .exists()
            )
            if not terms:
                messages.warning(self.request, "You must set a term to current.")
                return redirect("term")
        return super().form_valid(form)


class FuncaoChefiaDeleteView(LoginRequiredMixin, DeleteView):
    model = AcademicTerm
    success_url = reverse_lazy("terms")
    template_name = "corecode/core_confirm_delete.html"
    success_message = "The term {} has been deleted with all its attached content"

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.current == True:
            messages.warning(request, "Cannot delete term as it is set to current")
            return redirect("terms")
        messages.success(self.request, self.success_message.format(obj.name))
        return super(TermDeleteView, self).delete(request, *args, **kwargs)
    


class DepartamentoListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = Departamento
    template_name = "corecode/departamento_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AcademicTermForm()
        return context


class CategoriaListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = Categoria
    template_name = "corecode/categoria_list.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form"] = AcademicTermForm()
        return context
    


@login_required(login_url="/accounts/login/")
def funcao_chefia_add(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    funcao_chefia = FuncaoChefia(nome=request.POST["nome"])
    funcao_chefia.save()
    context = {"funcao_chefia":funcao_chefia,
               "abertura_actual":abertura}

    template_name = "employees/funcao_success.html"

    return render(request, template_name,context)



@login_required(login_url="/accounts/login/")
def funcao_chefia_list(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    
    chefias=FuncaoChefia.objects.filter(estado_objecto="activo")

    context ={'chefias':chefias,"abertura_actual":abertura}


    template_name = "corecode/funcao_chefia_list.html"
    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def funcao_chefia_edit(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    chefias = FuncaoChefia.objects.get(pk=id)
    context = {"chefias":chefias,"abertura_actual":abertura}
    template_name = "corecode/funcao_chefia_edit.html"
    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def funcao_chefia_detail(request, id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    chefias = FuncaoChefia.objects.get(pk=id)
    context = {"chefias":chefias,"abertura_actual":abertura}
    template_name = "corecode/funcao_chefia_detail.html"
    return render(request, template_name, context)



@login_required(login_url="/accounts/login/")
def departamento_add(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    departamento = Departamento(nome=request.POST["nome"])
    departamento.save()
    context = {"departamento":departamento,"abertura_actual":abertura}

    template_name = "employees/departamento_success.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def departamento_list(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    departamentos=Departamento.objects.filter(estado_objecto="activo")

    context ={'departamentos':departamentos,"abertura_actual":abertura}


    template_name = "corecode/departamento_list.html"
    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def departamento_edit(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    departamentos = Departamento.objects.get(pk=id)
    context = {"departamentos":departamentos,"abertura_actual":abertura}
    template_name = "corecode/departamento_edit.html"
    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def departamento_detail(request, id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    departamentos = Departamento.objects.get(pk=id)
    context = {"departamentos":departamentos,"abertura_actual":abertura}
    template_name = "corecode/departamento_detail.html"
    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def deleteDepartamento(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    departamento = Departamento.objects.get(pk=request.POST['id'])
    departamento.estado_objecto="Eliminado"
    departamento.save()
    departamentos  = Departamento.objects.filter(estado_objecto='activo')
    context = {"departamentos":departamentos,"error":"Departamento Eliminado Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/departamento_list.html"

    return render(request, template_name,context)

@login_required(login_url="/accounts/login/")
def deleteFuncaoChefia(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    chefia = FuncaoChefia.objects.get(pk=request.POST['id'])
    chefia.estado_objecto="Eliminado"
    chefia.save()
    chefias  = FuncaoChefia.objects.filter(estado_objecto='activo')
    context = {"chefias":chefias,"error":"Função de Chefia Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/funcao_chefia_list.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def funcao_chefia_delete(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    chefias = FuncaoChefia.objects.get(pk=id)
    context = {"chefias": chefias,"error":"Função de Chefia Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/funcao_chefia_delete.html"
    return render(request, template_name, context)


 
 

@login_required(login_url="/accounts/login/")
def categoria_add(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    categoria = Categoria(nome=request.POST["nome"])
    categoria.save()
    context = {"categoria":categoria,"abertura_actual":abertura}

    template_name = "employees/categoria_success.html"

    return render(request, template_name,context)

@login_required(login_url="/accounts/login/")
def categoria_nova_add(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    categoria_nova = CategoriaNova(nome=request.POST["nome"])
    categoria_nova.save()
    context = {"categoria_nova":categoria_nova,"abertura_actual":abertura}

    template_name = "employees/categoria_nova_success.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def categoria_list(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    categorias=Categoria.objects.all()

    context ={'categorias':categorias,"abertura_actual":abertura}

    template_name = "corecode/categoria_list.html"
    return render(request, template_name,context)

@login_required(login_url="/accounts/login/")
def categoria_nova_list(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    categorias_novas =CategoriaNova.objects.all()

    context ={'categorias_novas':categorias_novas,"abertura_actual":abertura}

    template_name = "corecode/categoria_nova_list.html"
    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def categoria_edit(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias = Categoria.objects.get(pk=id)
    context = {"categorias":categorias,"abertura_actual":abertura}
    template_name = "corecode/categoria_edit.html"
    return render(request, template_name, context)

@login_required(login_url="/accounts/login/")
def categoria_nova_edit(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias_novas = CategoriaNova.objects.get(pk=id)
    context = {"categorias_novas":categorias_novas,"abertura_actual":abertura}
    template_name = "corecode/categoria_nova_edit.html"
    return render(request, template_name, context)



@login_required(login_url="/accounts/login/")
def categoria_detail(request, id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias = Categoria.objects.get(pk=id)
    context = {"categorias":categorias,"abertura_actual":abertura}
    template_name = "corecode/categoria_detail.html"
    return render(request, template_name, context)

@login_required(login_url="/accounts/login/")
def categoria_nova_detail(request, id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias_novas = CategoriaNova.objects.get(pk=id)
    context = {"categorias_novas":categorias_novas,"abertura_actual":abertura}
    template_name = "corecode/categoria_nova_detail.html"
    return render(request, template_name, context)

@login_required(login_url="/accounts/login/")
def deleteCategoria(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categoria = Categoria.objects.get(pk=request.POST['id'])
    categoria.estado_objecto="Eliminado"
    categoria.save()
    categorias  = Categoria.objects.filter(estado_objecto='activo')
    context = {"categorias":categorias,"error":"Categoria Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/categoria_list.html"

    return render(request, template_name,context)

@login_required(login_url="/accounts/login/")
def deleteCategoriaNova(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categoria_nova = CategoriaNova.objects.get(pk=request.POST['id'])
    categoria_nova.estado_objecto="Eliminado"
    categoria_nova.save()
    categorias_novas  = CategoriaNova.objects.filter(estado_objecto='activo')
    context = {"categorias_novas":categorias_novas,"error":"Categoria Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/categoria_nova_list.html"

    return render(request, template_name,context)

@login_required(login_url="/accounts/login/")
def departamento_delete(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    departamentos = Departamento.objects.get(pk=id)
    context = {"departamentos": departamentos,"error":"Departamento Eliminado Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/departamento_delete.html"
    return render(request, template_name, context)

@login_required(login_url="/accounts/login/")
def categoria_delete(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias = Categoria.objects.get(pk=id)
    context = {"categorias": categorias,"error":"Categoria Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/categoria_delete.html"
    return render(request, template_name, context)

@login_required(login_url="/accounts/login/")
def categoria_nova_delete(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias_novas = CategoriaNova.objects.get(pk=id)
    context = {"categorias_novas": categorias_novas,"error":"Categoria Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/categoria_nova_delete.html"
    return render(request, template_name, context)




@login_required(login_url="/accounts/login/")
def edit_success(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    template_name = "corecode/edit_success.html"
    return render(request, template_name)



@login_required(login_url="/accounts/login/")
def editFuncaoChefia(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    chefias = FuncaoChefia.objects.get(pk=request.POST['id'])
    context = {"chefias":chefias}

    chefias.nome = request.POST['funcao_chefia']

    chefias.save()

    context = {"chefias":chefias,"abertura_actual":abertura}

    template_name = "corecode/edit_success.html"

    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def editDepartamento(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    departamentos = Departamento.objects.get(pk=request.POST['id'])
    context = {"departamentos":departamentos}

    departamentos.nome = request.POST['departamento']

    departamentos.save()

    context = {"departamentos":departamentos,"abertura_actual":abertura}

    template_name = "corecode/edit_success.html"

    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def editCategoria(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias = Categoria.objects.get(pk=request.POST['id'])
    context = {"categorias":categorias}

    categorias.nome = request.POST['categoria']

    categorias.save()

    context = {"categorias":categorias,"abertura_actual":abertura}

    template_name = "corecode/edit_success.html"

    return render(request, template_name, context)  

@login_required(login_url="/accounts/login/")
def editCategoriaNova(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    categorias_novas = CategoriaNova.objects.get(pk=request.POST['id'])
    context = {"categorias_novas":categorias_novas}

    categorias_novas.nome = request.POST['categoria_nova']

    categorias_novas.save()

    context = {"categorias_novas":categorias_novas,"abertura_actual":abertura}

    template_name = "corecode/edit_success.html"

    return render(request, template_name, context)  


@login_required(login_url="/accounts/login/")
def index(request):
    dias=0
    # Employee.objects.filter(currenr_stutus="activo").filter(estado_objecto="activo").count()
    funcionarios=Employee.objects.filter(current_status="activo").filter(estado_objecto="activo").count()
    #funcionarios_ref=Employee.objects.filter(estado_objecto="reformado").count()
    funcionarios_lic=Employee.objects.filter(current_status="licenciado").count()
    funcionarios_pre=Employee.objects.filter(current_status="pre-reformado").count()
    
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
    ultima_abertura= Abertura_Prova_Vida.objects.filter(estado_actual="fechada").order_by('data_de_abertura').last()
    #dias=0

    if len(abertura)>0:
        prova_vida = ProvaVida.objects.filter(abertura_prova_vida=abertura[0])
        #data_inicio = datetime.strptime(abertura[0].data_de_abertura, "%Y-%m-%d")
        #data_fim = datetime.strptime(abertura[0].data_de_fim, "%Y-%m-%d")
        

        # Calculate the difference in days
        difference = abertura[0].data_de_fim - abertura[0].data_de_abertura

        # Extract the difference in days   
        dias = difference.days

        abertura_actual=abertura[0]

    
        provas=len(prova_vida)
    else:
        provas=0
        abertura_actual=None


   
    
    if dias>1:
        num="dias"
    else:
        num="Dia"
    
    context = {"user":request.user,
        "qtd_func":funcionarios,
               "qtd_func_lic":funcionarios_lic,
               "qtd_func_pre":funcionarios_pre,
               "qtd_conc":provas,
               "qtd_pen":funcionarios-provas,
               "abertura_actual":abertura_actual,
               "ultima_abertura":ultima_abertura,
               "dias":"{} {}".format(dias,num)
               }
    template_name = "index.html"
    return render(request, template_name, context)


from django.shortcuts import render
from .forms import ExcelUploadForm
from apps.employees.models import Employee
from openpyxl import load_workbook

@login_required(login_url="/accounts/login/")
def upload_excel_categoria(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):

               

                Categoria.objects.create(
                    nome=row[0],
                   
                )
                context= {"msg":"Categorias Importadas com Sucesso!","tipo":"importação de categoria","abertura_actual":abertura}
            return render(request, 'corecode/categoria_success.html',context)

@login_required(login_url="/accounts/login/")
def upload_excel_categoria_nova(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):

               

                CategoriaNova.objects.create(
                    nome=row[0],
                   
                )
                context= {"msg":"Categorias Importadas com Sucesso!","tipo":"importação de categoria nova","abertura_actual":abertura}
            return render(request, 'corecode/categoria_nova_success.html',context)
        


@login_required(login_url="/accounts/login/")
def upload_excel_departamento(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):

               

                Departamento.objects.create(
                    nome=row[0],
                   
                )
            context= {"msg":"Departamentos Importado com Sucesso!","tipo":"importação de departamentos","abertura_actual":abertura}
            return render(request, 'employees/succes.html',context)
        


@login_required(login_url="/accounts/login/")
def upload_excel_funcao_chefia(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue


                FuncaoChefia.objects.create(
                    nome=row[0],
                   
                )

                context= {"msg":"Funções de Chefia Importada com Sucesso!","tipo":"importação de função de chefia","abertura_actual":abertura}
            return render(request, 'corecode/chefia_success.html',context)
        


@login_required(login_url="/accounts/login/")
def upload_excel_direcao(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):

               

                DirecaoAlocacao.objects.create(
                    nome=row[0],
                   
                )
            context= {"msg":"Direcções de Alocação Importada com Sucesso!","tipo":"importação de direcção de alocação","abertura_actual":abertura}
            return render(request, 'corecode/direccao_success.html',context)
        



@login_required(login_url="/accounts/login/")
def direccao_add(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    direccao = DirecaoAlocacao(nome=request.POST["nome"])
    direccao.save()
    context = {"direccao":direccao,"abertura_actual":abertura}

    template_name = "employees/direccao_success.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def direccao_list(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    direccoes=DirecaoAlocacao.objects.all()

    context ={'direccoes':direccoes,"abertura_actual":abertura}

    template_name = "corecode/direccao_list.html"
    return render(request, template_name,context)
 

@login_required(login_url="/accounts/login/")
def direccao_edit(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    direccoes = DirecaoAlocacao.objects.get(pk=id)
    context = {"direccoes":direccoes ,"abertura_actual":abertura}
    template_name = "corecode/direccao_edit.html"
    return render(request, template_name, context)



@login_required(login_url="/accounts/login/")
def direccao_detail(request, id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    direccoes = DirecaoAlocacao.objects.get(pk=id)
    context = {"direccoes":direccoes,"abertura_actual":abertura}
    template_name = "corecode/direccao_detail.html"
    return render(request, template_name, context)

@login_required(login_url="/accounts/login/")
def direccao_delete(request,id):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    direccoes = DirecaoAlocacao.objects.get(pk=id)
    context = {"direccoes": direccoes,"error":"Direcçã Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/direccao_delete.html"
    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def editDireccao(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    direccoes = DirecaoAlocacao.objects.get(pk=request.POST['id'])
    context = {"direccoes":direccoes,"abertura_actual":abertura}

    direccoes.nome = request.POST['direccao']

    direccoes.save()

    context = {"direccoes":direccoes}

    template_name = "corecode/edit_success.html"

    return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def deleteDireccao(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()
    direccao = DirecaoAlocacao.objects.get(pk=request.POST['id'])
    direccao.estado_objecto="Eliminado"
    direccao.save()
    direccoes  = DirecaoAlocacao.objects.filter(estado_objecto='activo')
    context = {"direccoes":direccoes,"error":"Direcção Eliminada Com Sucesso!","abertura_actual":abertura}
    template_name = "corecode/direccao_list.html"

    return render(request, template_name,context)


from django.contrib.auth import logout


@login_required(login_url="/accounts/login/")
def logout_view(request):

    logout(request)
    template_name = "registration/login.html"

    return render(request, template_name)




def categoria_list_ajax(request,page):
    print("I have been callii list categoria")


    categorias  = Categoria.objects.all()
 
    
    p = Paginator(categorias, 10)
    
   
    context = {"page_obj": p.get_page(page),
               'categorias':categorias,
            }

    serialized_data = []
    for categoria in  p.get_page(page):
        serialized_data.append({

            'id': categoria.id,
            'nome': categoria.nome,
    
            # Add other fields as needed
        })

    return JsonResponse({
        'count': p.count,
        'results': serialized_data,
    })

def categoria_nova_list_ajax(request,page):
    print("I have been callii list categoria")


    categorias_novas  = CategoriaNova.objects.all()
 
    
    p = Paginator(categorias_novas, 10)
    
   
    context = {"page_obj": p.get_page(page),
               'categorias_novas':categorias_novas,
            }

    serialized_data = []
    for categoria_nova in  p.get_page(page):
        serialized_data.append({

            'id': categoria_nova.id,
            'nome': categoria_nova.nome,
    
            # Add other fields as needed
        })

    return JsonResponse({
        'count': p.count,
        'results': serialized_data,
    })






        



        


        

        





