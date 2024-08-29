import csv
import datetime
from datetime import datetime 
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.forms import widgets
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy
from django.views.generic import DetailView, ListView, View
from django.shortcuts import HttpResponseRedirect, redirect, render
from django.views.generic.edit import CreateView, DeleteView, UpdateView  
from django.core.paginator import Paginator
import io
from openpyxl import Workbook
from apps.prova_vida.models import Abertura_Prova_Vida
from reportlab.lib.pagesizes import letter, landscape
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageTemplate, Frame, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import landscape, letter, A3
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer
from django.contrib.auth.models import User
from django.templatetags.static import static
#from apps.finance.models import Invoice


from .forms import ExcelUploadForm
from openpyxl import load_workbook
from django.db import transaction
import logging
from datetime import datetime
import math
from .models import Employee, EmployeeBulkUpload

from apps.corecode.models import Categoria, CategoriaNova, Citizenship,DirecaoAlocacao,FuncaoChefia

from .forms import (
    EmployeeForm 
)



class EmployeeListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = Employee
    # context_object_name = 'employee_list'   # your own name for the list as a template variable
    template_name = "employees/employee_list_actual.html" # own template name/location



    def get_context_data(self, **kwargs):
        context = super(EmployeeListView, self).get_context_data(**kwargs)
        # context["employees"] = Employee.objects.filter(current_status='active') # Get 5 employees with status 'active'
        context["employees"] = Employee.objects.all()
        return context


class EmployeeDetailView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = "employees/employee_detail.html"

    def get_context_data(self, **kwargs):
        context = super(EmployeeDetailView, self).get_context_data(**kwargs)
    #    context["payments"] = Invoice.objects.filter(employee=self.object)
        return context


class EmployeeCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Employee
    fields = "__all__"
    success_message = "Novo Funcionário Adicionado com sucesso."

    def get_form(self):
        """add date picker in forms"""
        form = super(EmployeeCreateView, self).get_form()
        form.fields["date_of_birth"].widget = widgets.DateInput(attrs={"type": "date"})
        form.fields["address"].widget = widgets.Textarea(attrs={"rows": 2})
        #form.fields["others"].widget = widgets.Textarea(attrs={"rows": 2})
        return form


class EmployeeUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Employee
    fields = "__all__"
    success_message = "Fucnionário actualizado com sucesso."

    def get_form(self):
        """add date picker in forms"""
        form = super(EmployeeUpdateView, self).get_form()
        form.fields["date_of_birth"].widget = widgets.DateInput(attrs={"type": "date"})
        form.fields["date_of_employment"].widget = widgets.DateInput(
            attrs={"type": "date"}
        )
        form.fields["address"].widget = widgets.Textarea(attrs={"rows": 2})
        form.fields["others"].widget = widgets.Textarea(attrs={"rows": 2})
        form.fields['photo'].widget = widgets.FileInput()
        return form


class EmployeeDeleteView(LoginRequiredMixin, DeleteView):
    model = Employee
    success_url = reverse_lazy("employee-list")


class EmployeeBulkUploadView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = EmployeeBulkUpload
    template_name = "employees/employee_upload.html"
    fields = ["csv_file"]
    success_url = "/employee/list"
    success_message = "Успешно загруженные сотрудники"


class DownloadCSVViewdownloadcsv(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="employee_template.csv"'

        writer = csv.writer(response)
        writer.writerow(
            [
                "registration_number",
                "surname",
                "firstname",
                "other_names",
                "gender",
                "parent_number",
                "address",
                "current_class",
            ]
        )

        return response



@login_required(login_url="/accounts/login/")
def addFuncionario(request):
    
    try:
        direcao = DirecaoAlocacao.objects.get(pk=request.POST['direcao_alocacao'])
    except:
        direcao=None

    try:
        categoria = Categoria.objects.get(pk=request.POST['categoria'])
    except:
            categoria=None

    try:
        funcao_chefia = FuncaoChefia.objects.get(pk=request.POST['funcoes_chefias_nova'])
    except:
            funcao_chefia=None


    try:
        funcao_chefia_antiga = CategoriaNova.objects.get(pk=request.POST['funcoes_chefias_nova'])
    except:
            funcao_chefia_antiga=None        

    try:
        correio_electronico = request.POST['correio_electronico']
    except:
            correio_electronico=None

    try:
        area_de_formacao = request.POST['area_de_formacao']
    except:
            area_de_formacao=None

    try:
        numero_telefone = request.POST['telefone']
    except:
            numero_telefone=None

    try:
        estado_civil = request.POST['estado_civil']
    except:
            estado_civil=None


    try:
        reforma = request.POST['reforma']
    except:
            reforma=None   

    try:
        nacionality = request.POST['nacionalidade']
    except:
            nacionality=None   
    try:
        categoria_antiga_id = request.POST['categoria_antiga']
        categoria_antiga = Categoria.objects.get(id=categoria_antiga_id)
    except:
            categoria_antiga=None   
    try:
        categoria_nova_id = request.POST['categoria_antiga']
        categoria_nova = Categoria.objects.get(id=categoria_nova_id)
    except:
            categoria_nova=None   
    

    data_admissao = request.POST.get('data_admissao')
    data_admissao = request.POST.get('data_de_emissao')

    tempo_na_empresa = datetime.now()
    if data_admissao:
        data_admissao = datetime.strptime(data_admissao, '%Y-%m-%d')
        data_atual = datetime.now()
        tempo_na_empresa = data_atual - data_admissao
            
    anos_na_empresa = math.floor(tempo_na_empresa.days / 365)  
    diuturnidade = 0

    funcionario=Employee(      
        # Dados Pessoais   
        numero_mecanografico = request.POST['numero_mecanografico'],
        firstname=request.POST['primeiro_nome'],
        current_status = 'activo',
        personnel_number = request.POST['numero_bi'],
        date_of_birth = datetime.strptime(request.POST.get('data_nascimento', ''), '%Y-%m-%d') if request.POST.get('data_nascimento') else None,
        data_de_emissao = datetime.strptime(request.POST.get('data_de_emissao', ''), '%Y-%m-%d') if request.POST.get('data_de_emissao') else None,
        data_de_validade = datetime.strptime(request.POST.get('data_de_validade', ''), '%Y-%m-%d') if request.POST.get('data_de_validade') else None,
        numero_seguranca_social = request.POST['numero_seguranca_social'],
        nacionalidade = request.POST['nacionalidade'],
        provincia_nascimento = request.POST['provincia_nascimento'],
        gender = request.POST['genero'],
        estado_civil = estado_civil,
        numero_dependentes = request.POST['numero_dependentes'],
        morada = request.POST['morada'],
        provincia_residencia = request.POST['provincia_residencia'],
        telefone = numero_telefone,
        correio_electronico = correio_electronico,
        
        # Dados Profissionais
        data_de_admissao = datetime.strptime(request.POST.get('data_admissao', ''), '%Y-%m-%d') if request.POST.get('data_admissao') else None,
        direccao= direcao,
        funcao_chefia = funcao_chefia,
        categoria_laboral_antiga = categoria_antiga,
        categoria_laboral_nova=categoria_nova,
        vencimento_mensal = request.POST['vencimento_mensal'],
        habilitacao = request.POST['habilitacao_literaria'],
        area_de_formacao = area_de_formacao,
        reforma = reforma,
        tempo_na_empresa = anos_na_empresa,
    )

    funcionario.save()

    context = {"funcionario":funcionario}

    template_name = "employees/funcionario_success.html"

    return render(request, template_name,context)



   
 

@login_required(login_url="/accounts/login/")
def employee_list(request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta").first()

    
    funcionarios  = Employee.objects.filter(estado_objecto='activo').order_by("firstname")
    categorias_antigas  = Categoria.objects.filter(tipo="antiga")
    categorias_nova  = Categoria.objects.filter(tipo="nova")
    direccoes  = DirecaoAlocacao.objects.all()
    funcoes_chefia  = FuncaoChefia.objects.all()
    p = Paginator(funcionarios, 100)
    
   
    context = {"funcionarios": funcionarios,
               'categorias_antigas':categorias_antigas,
               'categorias_novas':categorias_nova,
               'direccoes':direccoes,
               'funcoes_chefia':funcoes_chefia,
               'abertura_actual':abertura,}

    template_name = "employees/employee_list.html"
    return render(request, template_name,context)



@login_required(login_url="/accounts/login/")
def employee_list_type(request,tipo):


    
    funcionarios  = Employee.objects.filter(current_status=tipo)
    categorias  = Categoria.objects.all()
    direcoes  = DirecaoAlocacao.objects.all()
    funcoes_chefia  = FuncaoChefia.objects.all()
    
   
    context = {"funcionarios": funcionarios,
               'categorias':categorias,
               'direcoes':direcoes,
               'funcoes_chefia':funcoes_chefia,}

    template_name = "employees/employee_list.html"
    return render(request, template_name,context)



@login_required(login_url="/accounts/login/")
def employee_edit(request,id):
      
    categorias_antigas  = Categoria.objects.filter(tipo="antiga")
    categorias_novas  = Categoria.objects.filter(tipo="nova")
    direccoes  = DirecaoAlocacao.objects.filter(estado_objecto='activo')
    funcoes_chefia  = FuncaoChefia.objects.filter(estado_objecto='activo')
   
    
    funcionario = Employee.objects.get(pk=id)

    date_of_birth = request.POST.get('date_of_birth')
    if date_of_birth:
        funcionario.date_of_birth = datetime.strptime(date_of_birth, '%Y-%m-%d')

    data_de_admissao = request.POST.get('data_de_admissao')
    if data_de_admissao:
        funcionario.data_de_admissao = datetime.strptime(data_de_admissao, '%Y-%m-%d')

    data_de_demissao = request.POST.get('data_de_demissao')
    if data_de_demissao:
        funcionario.data_de_demissao = datetime.strptime(data_de_demissao, '%Y-%m-%d')

    data_de_validade = request.POST.get('data_de_validade')
    if data_de_validade:
        funcionario.data_de_validade = datetime.strptime(data_de_validade, '%Y-%m-%d')

    # funcionario.data_de_admissao=funcionario.data_de_admissao.strftime('%Y-%m-%d')
    # funcionario.date_of_birth=funcionario.date_of_birth.strftime('%Y-%m-%d')
    # funcionario.data_de_demissao=funcionario.data_de_demissao.strftime('%Y-%m-%d')
    # funcionario.data_de_validade=funcionario.data_de_validade.strftime('%Y-%m-%d')
    #print(funcionario.data_de_admissao)
    #print(funcionario.date_of_birth)
    context = {"funcionario": funcionario,
            'categorias_antigas':categorias_antigas,
            'categorias_novas':categorias_novas,
            'direccoes':direccoes,
            'funcoes_chefia':funcoes_chefia,}
    template_name = "employees/employee_edit.html"
    return render(request, template_name, context)



@login_required(login_url="/accounts/login/")
def employee_delete(request,id):
    funcionario = Employee.objects.get(pk=id)
    context = {"funcionario": funcionario,}
    template_name = "employees/employee_delete.html"
    return render(request, template_name, context)





@login_required(login_url="/accounts/login/")
def employee_detail(request,id):
    funcionario = Employee.objects.get(pk=id)
    context = {"funcionario":funcionario}
    template_name = "employees/employee_detail.html"
    return render(request, template_name, context)



@login_required(login_url="/accounts/login/")
def editFuncionario(request):
    

    funcionario = Employee.objects.get(pk=request.POST.get('id'))

    try:
        direcao = DirecaoAlocacao.objects.get(pk=request.POST['direcao_alocacao'])
    except:
        direcao=None

    try:
        funcao_chefia = FuncaoChefia.objects.get(pk=request.POST['funcoes_chefias_nova'])
    except:
        funcao_chefia=None

    try:
        funcao_chefia_antiga = CategoriaNova.objects.get(pk=request.POST['funcoes_chefias_nova'])
    except:
            funcao_chefia_antiga=None        

    try:
        correio_electronico = request.POST['correio_electronico']
    except:
            correio_electronico=None

    try:
        area_de_formacao = request.POST['area_de_formacao']
    except:
            area_de_formacao=None

    try:
        telefone = request.POST['telefone']
    except:
            telefone=None

    try:
        estado_civil = request.POST['estado_civil']
    except:
            estado_civil=None
    try:
        reforma = request.POST['reforma']
    except:
            reforma=None    
    try:
        categoria_antiga_id = request.POST['categoria_antiga']
        categoria_antiga = Categoria.objects.get(id=categoria_antiga_id)
    except:
            categoria_antiga=None   
    try:
        categoria_nova_id = request.POST['categoria_nova']
        categoria_nova = Categoria.objects.get(id=categoria_nova_id)
    except:
            categoria_nova=None   
    

    data_admissao = request.POST.get('data_admissao')
    data_admissao = request.POST.get('data_de_emissao')

    # Verificação e conversão de datas
    data_de_demissao = request.POST.get('data_de_demissao')
    if data_de_demissao:
        funcionario.data_de_demissao = datetime.strptime(data_de_demissao, '%Y-%m-%d')

    data_de_validade = request.POST.get('data_de_validade')
    if data_de_validade:
        funcionario.data_de_validade = datetime.strptime(data_de_validade, '%Y-%m-%d')
    
    data_nascimento = request.POST.get('data_nascimento')
    if data_nascimento:
        funcionario.date_of_birth = datetime.strptime(data_nascimento, '%Y-%m-%d')

    data_admissao = request.POST.get('data_admissao')
    if data_admissao:
        funcionario.data_de_admissao = datetime.strptime(data_admissao, '%Y-%m-%d')

    data_demissao = request.POST.get('data_de_demissao')
    data_admissao = request.POST.get('data_admissao')
    if data_admissao:
        data_admissao = datetime.strptime(data_admissao, '%Y-%m-%d')

    if data_demissao:
        data_demissao = datetime.strptime(data_demissao, '%Y-%m-%d')
        tempo_na_empresa = data_demissao - data_admissao
    else:
        data_atual = datetime.now()
        tempo_na_empresa = data_atual - data_admissao
            
    anos_na_empresa = math.floor(tempo_na_empresa.days / 365)  

    # Dados Pessoais 
    funcionario.numero_mecanografico = request.POST['numero_mecanografico']
    funcionario.firstname = request.POST['primeiro_nome']
    funcionario.personnel_number = request.POST['numero_bi']
    funcionario.date_of_birth = datetime.strptime(request.POST.get('data_nascimento', ''), '%Y-%m-%d') if request.POST.get('data_nascimento') else None
    funcionario.data_de_emissao = datetime.strptime(request.POST.get('data_de_emissao', ''), '%Y-%m-%d') if request.POST.get('data_de_emissao') else None
    funcionario.data_de_validade = datetime.strptime(request.POST.get('data_de_validade', ''), '%Y-%m-%d') if request.POST.get('data_de_validade') else None
    funcionario.numero_seguranca_social = request.POST['numero_seguranca_social']
    funcionario.nacionalidade = request.POST['nacionalidade']
    funcionario.provincia_nascimento = request.POST['provincia_nascimento']
    funcionario.gender = request.POST['genero']
    funcionario.estado_civil = estado_civil
    funcionario.numero_dependentes = request.POST['numero_dependentes']
    funcionario.morada = request.POST['morada']
    funcionario.provincia_residencia = request.POST['provincia_residencia']
    funcionario.telefone = telefone
    funcionario.correio_electronico = correio_electronico

    funcionario.tempo_na_empresa = anos_na_empresa
    
    # Dados Profissionais
    funcionario.habilitacao = request.POST.get('habilitacao_literaria')
    funcionario.reforma = reforma
    funcionario.data_de_admissao = data_admissao
    funcionario.funcao_chefia = funcao_chefia
    funcionario.direccao = direcao
    funcionario.categoria_laboral_antiga = categoria_antiga
    funcionario.categoria_laboral_nova=categoria_nova
    funcionario.vencimento_mensal = request.POST.get('vencimento_mensal')
    funcionario.area_de_formacao = area_de_formacao

    data_fim_contrato = request.POST.get('data_fim_contrato')
    if data_fim_contrato:
        funcionario.data_fim_contrato = datetime.strptime(data_fim_contrato, '%Y-%m-%d')

    funcionario.current_status = 'activo'
    funcionario.origem = 'incluso'    

    funcionario.save()

    context = {"funcionario": funcionario}
    template_name = "corecode/edit_success.html"

    return render(request, template_name, context)



@login_required(login_url="/accounts/login/")
def deleteFuncionario(request):
    funcionario = Employee.objects.get(pk=request.POST['id'])
    funcionario.estado_objecto="Eliminado"
    funcionario.save()
    funcionarios  = Employee.objects.filter(estado_objecto='activo')
    context = {"funcionarios":funcionarios,"error":"Funcionário Eliminado com sucesso"}
    template_name = "employees/employee_list.html"

    return render(request, template_name,context)




from .forms import ExcelUploadForm
from .models import Employee
from openpyxl import load_workbook

@login_required(login_url="/accounts/login/")
def upload_excel(request):
    errors = []
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                numero_mecanografico = row[0]
                firstname = row[1]
                vencimento_mensal = row[2]
                personnel_number = row[3]
                date_of_birth = row[4]
                data_de_emissao = row[5]
                data_de_validade = row[6]
                numero_seguranca_social = row[7]
                nacionalidade = row[8]
                provincia_nascimento = row[9]
                gender = row[10]
                estado_civil = row[11]
                numero_dependentes = row[12]
                morada = row[13]
                provincia_residencia = row[14]
                telefone = row[15]
                correio_electronico = row[16]
                data_de_admissao = row[17]
                direcao_nome = row[18]
                funcao_chefia = row[19]
                categoria_antiga_nome = row[20]
                categoria_nova_nome = row[21]
                habilitacao = row[22]
                area_de_formacao = row[23]
                reforma = row[24]
                tempo_na_empresa = row[25]

                # Conversão de datas
                def parse_date(date_value):
                    if isinstance(date_value, str):
                        try:
                            return datetime.strptime(date_value, "%Y-%m-%d").date()
                        except ValueError:
                            errors.append(f"Formato de data inválido: {date_value}")
                            return None
                    elif isinstance(date_value, datetime):
                        return date_value.date()
                    return None

                data_admissao = parse_date(data_de_admissao)
                data_de_emissao = parse_date(data_de_emissao)
                data_de_validade = parse_date(data_de_validade)
                data_de_nascimento = parse_date(date_of_birth)

                # Validação de campos
                def validate_field(value):
                    return value if value not in [None, 'none', 'None', 0] else None

                categoria_antiga = Categoria.objects.filter(nome=categoria_antiga_nome, tipo='antiga').first()
                categoria_nova = Categoria.objects.filter(nome=categoria_nova_nome, tipo='nova').first()
                funcao_chefia = FuncaoChefia.objects.filter(nome=funcao_chefia).first()
                direcao = DirecaoAlocacao.objects.filter(nome=direcao_nome).first()
                habilitacao = validate_field(habilitacao)
                firstname = validate_field(firstname)
                tempo_na_empresa = tempo_na_empresa

                try:
                    Employee.objects.create(
                        # Dados Pessoais   
                        numero_mecanografico = numero_mecanografico,
                        firstname=firstname,
                        personnel_number = personnel_number,
                        date_of_birth = data_de_nascimento,
                        data_de_emissao = data_de_emissao,
                        data_de_validade = data_de_validade,
                        numero_seguranca_social = numero_seguranca_social,
                        nacionalidade = nacionalidade,
                        provincia_nascimento = provincia_nascimento,
                        gender = gender,
                        estado_civil = estado_civil,
                        numero_dependentes = numero_dependentes,
                        morada = morada,
                        provincia_residencia = provincia_residencia,
                        telefone = telefone,
                        correio_electronico = correio_electronico,
                        
                        # Dados Profissionais
                        data_de_admissao = data_admissao,
                        direccao= direcao,
                        funcao_chefia = funcao_chefia,
                        categoria_laboral_antiga = categoria_antiga,
                        categoria_laboral_nova=categoria_nova,
                        vencimento_mensal = vencimento_mensal,
                        habilitacao = habilitacao,
                        area_de_formacao = area_de_formacao,
                        reforma = reforma,
                        tempo_na_empresa = tempo_na_empresa,

                        current_status='activo',
                        origem='Primavera',
                        estado='Activo',
                    )
                except Exception as e:
                    errors.append(f"Erro ao importar funcionário: {str(e)}")
                    continue

            return render(request, 'employees/succes.html',{'errors':errors,"msg":"funcionario importado com sucesso","tipo":"importação de funcionário"})

    return render(request, 'employees/upload.html', {'form': form})

        #  return render(request, 'employees/succes.html',{'errors':errors,"msg":"funcionario importado com sucesso","tipo":"importação de funcionário"})
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'employees/upload.html', {'form': form})
# def upload_excel(request):
#     errors = []
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES['excel_file']
#             wb = load_workbook(excel_file)
#             ws = wb.active

#             with transaction.atomic():
#                 for row in ws.iter_rows(min_row=2, values_only=True):
#                     try:
#                         # Processa os dados da linha
#                         numero_mecanografico = row[14]
#                         numero_seguranca_social = row[15]
#                         firstname = row[0]
#                         gender = row[1]
#                         estado_civil = row[2]
#                         telefone = str(row[3]) if row[3] else None
#                         correio_electronico = str(row[4]) if row[4] else None
#                         personnel_number = row[5]
#                         data_de_emissao = row[6]
#                         data_de_validade = row[7]
#                         data_nascimento = row[8]
#                         localidade = row[9]
#                         morada = row[10]
#                         provincia_residencia = row[11]
#                         provincia_nascimento = row[12]
#                         naturalidade = row[13]
#                         # numero_beneficiario = row[15]
#                         data_admissao = row[16]
#                         direcao = DirecaoAlocacao.objects.filter(nome=row[17]).first()
#                         categoria = Categoria.objects.filter(nome=row[18]).first()
#                         vencimento_mensal = row[19]
#                         funcao_chefia = FuncaoChefia.objects.filter(nome=row[20]).first()
#                         numero_dependentes = row[21]
#                         current_status = row[22]
#                         habilitacao = str(row[23]) if row[23] else None
#                         area_de_formacao = str(row[24]) if row[24] else None
#                         estabelecimento = row[25]

#                         # Verifica e converte as datas
#                         if isinstance(data_admissao, datetime):
#                             data_admissao = data_admissao.date()
#                         if isinstance(data_nascimento, datetime):
#                             data_nascimento = data_nascimento.date()

#                         # Cria o registro do funcionário
#                         Employee.objects.create(
#                             numero_mecanografico=numero_mecanografico,
#                             numero_seguranca_social=numero_seguranca_social,
#                             firstname=firstname,
#                             gender=gender,
#                             estado_civil=estado_civil,
#                             telefone=telefone,
#                             correio_electronico=correio_electronico,
#                             personnel_number=personnel_number,
#                             data_de_emissao=data_de_emissao,
#                             data_de_validade=data_de_validade,
#                             date_of_birth=data_nascimento,
#                             # numero_beneficiario=numero_beneficiario,
#                             localidade=localidade,
#                             morada=morada,
#                             provincia_residencia=provincia_residencia,
#                             provincia_nascimento=provincia_nascimento,
#                             naturalidade=naturalidade,
#                             data_admissao=data_admissao,
#                             direccao=direcao,
#                             categoria_laboral=categoria,
#                             funcao_chefia=funcao_chefia,
#                             vencimento_mensal=vencimento_mensal,
#                             numero_dependentes=numero_dependentes,
#                             current_status=current_status,
#                             habilitacao=habilitacao,
#                             area_de_formacao=area_de_formacao,
#                             estabelecimento=estabelecimento,
#                             origem="Primavera",
#                         )
                        
#                     except Exception as e:
#                         errors.append(f"Erro na linha {row[0]}: {e}")
#                         continue

#             return render(request, 'employees/succes.html',{'errors':errors,"msg":"funcionario importado com sucesso","tipo":"importação de funcionário"})
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'employees/upload.html', {'form': form})

# def upload_excel(request):
#     errors = []
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES['excel_file']
#             wb = load_workbook(excel_file)
#             ws = wb.active

#             employees = []

#             with transaction.atomic():
#                 for row in ws.iter_rows(min_row=2, values_only=True):
#                     try:
#                         # Processa os dados da linha
#                         numero_mecanografico = row[14]
#                         numero_seguranca_social = row[15]
#                         firstname = row[0]
#                         gender = row[1]
#                         estado_civil = row[2]
#                         telefone = str(row[3]) if row[3] else None
#                         correio_electronico = str(row[4]) if row[4] else None
#                         personnel_number = row[5]
#                         data_de_emissao = row[6]
#                         data_de_validade = row[7]
#                         data_nascimento = row[8]
#                         localidade = row[9]
#                         morada = row[10]
#                         provincia_residencia = row[11]
#                         provincia_nascimento = row[12]
#                         naturalidade = row[13]
#                         data_admissao = row[16]
#                         direcao = DirecaoAlocacao.objects.filter(nome=row[17]).first()
#                         categoria = Categoria.objects.filter(nome=row[18]).first()
#                         vencimento_mensal = row[19]
#                         funcao_chefia = FuncaoChefia.objects.filter(nome=row[20]).first()
#                         numero_dependentes = row[21]
#                         current_status = row[22]
#                         habilitacao = str(row[23]) if row[23] else None
#                         area_de_formacao = str(row[24]) if row[24] else None
#                         estabelecimento = row[25]

#                         # Verifica e converte as datas
#                         if isinstance(data_admissao, datetime):
#                             data_admissao = data_admissao.date()
#                         if isinstance(data_nascimento, datetime):
#                             data_nascimento = data_nascimento.date()

#                         # Adiciona o funcionário à lista
#                         employees.append(Employee(
#                             numero_mecanografico=numero_mecanografico,
#                             numero_seguranca_social=numero_seguranca_social,
#                             firstname=firstname,
#                             gender=gender,
#                             estado_civil=estado_civil,
#                             telefone=telefone,
#                             correio_electronico=correio_electronico,
#                             personnel_number=personnel_number,
#                             data_de_emissao=data_de_emissao,
#                             data_de_validade=data_de_validade,
#                             date_of_birth=data_nascimento,
#                             localidade=localidade,
#                             morada=morada,
#                             provincia_residencia=provincia_residencia,
#                             provincia_nascimento=provincia_nascimento,
#                             naturalidade=naturalidade,
#                             data_admissao=data_admissao,
#                             direccao=direcao,
#                             categoria_laboral=categoria,
#                             funcao_chefia=funcao_chefia,
#                             vencimento_mensal=vencimento_mensal,
#                             numero_dependentes=numero_dependentes,
#                             current_status=current_status,
#                             habilitacao=habilitacao,
#                             area_de_formacao=area_de_formacao,
#                             estabelecimento=estabelecimento,
#                             origem="Primavera",
#                         ))

#                     except Exception as e:
#                         errors.append(f"Erro na linha com firstname {row[0]}: {e}")
#                         continue

#                 if employees:
#                     Employee.objects.bulk_create(employees)

#             if errors:
#                 return render(request, 'employees/upload.html', {'form': form, 'errors': errors})
#             else:
#                 return render(request, 'employees/success.html', {"msg": "Funcionários importados com"})


from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from .models import Employee
import io

left_image_path = "./static/imagens/core/left.jpg"  # Path to the left image file in static/imagens
right_image_path = "./static/imagens/core/right.jpg"

@login_required(login_url="/accounts/login/")    
def gerarPDF(request):

    data1 = [
        ['Bold Text', 'Italic Text', 'Normal Text'],
        ['Bold and Italic Text', 'Bold and Italic Text', 'Italic Text'],
        ['Normal Text', 'Bold Text', 'Italic Text']
    ]

    # Define styles
    styles = getSampleStyleSheet()
    bold_style = styles["BodyText"]
    bold_style.fontName = "Helvetica-Bold"
    italic_style = styles["BodyText"]
    italic_style.fontName = "Helvetica-Oblique"
    # Fetch employee data from the database
    employees = funcionarios  = Employee.objects.filter(estado_objecto='activo').order_by("firstname")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=lista_de_funcionarios_activos.pdf'

    doc = SimpleDocTemplate(response, pagesize=A3)
    elements = []

    # Add image at top center
    right_image_path = "./static/imagens/core/left.jpg"
    image_path = "./static/imagens/core/right.jpg"  # Replace with the path to your image
    image = Image(image_path, width=70, height=70)
    image1 = Image(right_image_path, width=200, height=100)
    image.hAlign = 'LEFT'
    image1.hAlign = 'RIGHT'
    #doc.drawString(30,750,'OFFICIAL COMMUNIQUE')
    p=Paragraph("LISTA TOTAL DE FUNCIONÁRIOS",bold_style)
    #p.hAlign='RIGHT'
    #elements.append(image
    elements.append(image1)
    elements.append(p)
   
    elements.append(Spacer(1, 12))  # Add some space between image and table

    # Table data
    #queryset = YourModel.objects.all()  # Replace 'YourModel' with the name of your Django model
    data = [["Nome", "Nº BI", "Nº Mecanográfico", "Género", "Data de Nascimento", "Estado"]]  # Table headers

    for employee in employees:
        data.append([employee.firstname,
            employee.personnel_number,
            employee.numero_mecanografico,
            employee.gender,
            str(employee.date_of_birth),
            employee.estado_pv])
    # Create table

    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(table)

    doc.build(elements)

    return response

def gerarLicenciadosPDF(request):

    data1 = [
        ['Bold Text', 'Italic Text', 'Normal Text'],
        ['Bold and Italic Text', 'Bold and Italic Text', 'Italic Text'],
        ['Normal Text', 'Bold Text', 'Italic Text']
    ]

    # Define styles
    styles = getSampleStyleSheet()
    bold_style = styles["BodyText"]
    bold_style.fontName = "Helvetica-Bold"
    italic_style = styles["BodyText"]
    italic_style.fontName = "Helvetica-Oblique"
    # Fetch employee data from the database
    employees = funcionarios  = Employee.objects.filter(current_status='licenciado').order_by("firstname")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=lista_de_funcionarios_licenciados.pdf'

    doc = SimpleDocTemplate(response, pagesize=A3)
    elements = []

    # Add image at top center
    right_image_path = "./static/imagens/core/left.jpg"
    image_path = "./static/imagens/core/right.jpg"  # Replace with the path to your image
    image = Image(image_path, width=70, height=70)
    image1 = Image(right_image_path, width=200, height=100)
    image.hAlign = 'LEFT'
    image1.hAlign = 'RIGHT'
    #doc.drawString(30,750,'OFFICIAL COMMUNIQUE')
    p=Paragraph("LISTA TOTAL DE FUNCIONÁRIOS LICENCIADOS",bold_style)
    #p.hAlign='RIGHT'
    #elements.append(image
    elements.append(image1)
    elements.append(p)
   
    elements.append(Spacer(1, 12))  # Add some space between image and table

    # Table data
    #queryset = YourModel.objects.all()  # Replace 'YourModel' with the name of your Django model
    data = [["Nome", "Nº BI", "Nº Mecanográfico", "Género", "Data de Nascimento", "Estado"]]  # Table headers

    for employee in employees:
        data.append([employee.firstname,
            employee.personnel_number,
            employee.numero_mecanografico,
            employee.gender,
            str(employee.date_of_birth),
            employee.estado_pv])
    # Create table

    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(table)

    doc.build(elements)

    return response

def gerarPreReformaPDF(request):

    data1 = [
        ['Bold Text', 'Italic Text', 'Normal Text'],
        ['Bold and Italic Text', 'Bold and Italic Text', 'Italic Text'],
        ['Normal Text', 'Bold Text', 'Italic Text']
    ]

    # Define styles
    styles = getSampleStyleSheet()
    bold_style = styles["BodyText"]
    bold_style.fontName = "Helvetica-Bold"
    italic_style = styles["BodyText"]
    italic_style.fontName = "Helvetica-Oblique"
    # Fetch employee data from the database
    employees = funcionarios  = Employee.objects.filter(current_status='pre-reformado').order_by("firstname")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=lista_de_funcionarios_pre_reformados.pdf'

    doc = SimpleDocTemplate(response, pagesize=A3)
    elements = []

    # Add image at top center
    right_image_path = "./static/imagens/core/left.jpg"
    image_path = "./static/imagens/core/right.jpg"  # Replace with the path to your image
    image = Image(image_path, width=70, height=70)
    image1 = Image(right_image_path, width=200, height=100)
    image.hAlign = 'LEFT'
    image1.hAlign = 'RIGHT'
    #doc.drawString(30,750,'OFFICIAL COMMUNIQUE')
    p=Paragraph("LISTA TOTAL DE FUNCIONÁRIOS PRÉ-REFORMADOS",bold_style)
    #p.hAlign='RIGHT'
    #elements.append(image
    elements.append(image1)
    elements.append(p)
   
    elements.append(Spacer(1, 12))  # Add some space between image and table

    # Table data
    #queryset = YourModel.objects.all()  # Replace 'YourModel' with the name of your Django model
    data = [["Nome", "Nº BI", "Nº Mecanográfico", "Género", "Data de Nascimento", "Estado"]]  # Table headers

    for employee in employees:
        data.append([employee.firstname,
            employee.personnel_number,
            employee.numero_mecanografico,
            employee.gender,
            str(employee.date_of_birth),
            employee.estado_pv])
    # Create table

    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(table)

    doc.build(elements)

    return response

@login_required(login_url="/accounts/login/")
def get_employee_report(request):
    
    funcionarios_a  = Employee.objects.filter(current_status='activo').filter(estado_objecto='activo')
    funcionarios_r  = Employee.objects.filter(current_status='reformado').filter(estado_objecto='activo')
    funcionarios_p  = Employee.objects.filter(current_status='pre-reformado').filter(estado_objecto='activo')
    funcionarios_l  = Employee.objects.filter(current_status='licenciado').filter(estado_objecto='activo')

   
    context = {"funcionarios_a": funcionarios_a,
               "funcionarios_r": funcionarios_r,
               "funcionarios_p": funcionarios_p,
               "funcionarios_l": funcionarios_l,
               "qtd_func":funcionarios_a.count(),  
               "qtd_func_lic":funcionarios_l.count(),
               "qtd_func_pre":funcionarios_p.count(),
          }

    template_name = "employees/employee_report.html"
    return render(request, template_name,context)



@login_required(login_url="/accounts/login/")
def employee_list_pv(request,tipo):


    
    funcionarios  = Employee.objects.filter(estado_pv=tipo)
    categorias  = Categoria.objects.all()
    direcoes  = DirecaoAlocacao.objects.all()
    funcoes_chefia  = FuncaoChefia.objects.all()
    
   
    context = {"funcionarios": funcionarios,
               'categorias':categorias,
               'direcoes':direcoes,
               'funcoes_chefia':funcoes_chefia,}

    template_name = "employees/employee_list.html"
    return render(request, template_name,context)




def employee_list_ajax(request,page):
    print("I have been callii list employeee")

    
    funcionarios  = Employee.objects.filter(estado_objecto='activo').order_by("firstname")
    categorias  = Categoria.objects.all()
    direccoes  = DirecaoAlocacao.objects.all()
    funcoes_chefia  = FuncaoChefia.objects.all()
    
    p = Paginator(funcionarios, 100)
    
   
    context = {"page_obj": p.get_page(page),
               'categorias':categorias,
               'direccoes':direccoes,
               'funcoes_chefia':funcoes_chefia,}

    serialized_data = []
    for employee in  p.get_page(page):
        serialized_data.append({

            'id': employee.id,
            'firstname': employee.firstname,
            'surname': employee.surname,
            'date_of_birth':employee.date_of_birth,
            'personnel_number':employee.personnel_number,
            'estado_civil':employee.estado_civil,
            'telefone':employee.telefone,
            'localidade':employee.localidade,
            'numero_dependentes':employee.numero_dependentes,
            'morada':employee.morada,
            'numero_mecanografico':employee.numero_mecanografico,
            # Add other fields as needed
        })

    return JsonResponse({
        'count': p.count,
        'results': serialized_data,
    })


def export_to_excel(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Número Mecanográfico", 
    "Número INSS", 
    # "Novo Número Mecanográfico", 
    "Nome Completo", 
    "Gênero", 
    "Data de Nascimento", 
    "Nº do BI", 
    "Data de Emissão", 
    "Data de Validade",  
    "Estado Cívil", 
    "Localidade", 
    "Morada", 
    "Província de Residência",
    "Província de Nascimento", 
    "Naturalidade", 
    "Contacto Telefônico",
    "Correio Electrônico",  
    "Data de Admissão", 
    "Departamento", 
    "Estabelecimento", 
    "Categoria Laboral", 
    # "Nova Categoria Laboral", 
    "Função de Chefia", 
    #"Número de Dependentes", 
    "Vencimento Mensal", 
    # "Novo Vencimento Mensal", 
    "Situação", 
    # "Estado", 
    "Habilitação Literária",
    "Área de Formação"
    ]) 
    # "Estado da Prova de Vida", 
    # "Data de Realização da Prova de Vida", 
    # "Observações"   # Adding headers

    queryset = Employee.objects.all()  # Replace 'YourModel' with the name of your Django model

    for obj in queryset:
        categoria = ''
        funcao_chefia = ''
        direccao = ''
        
        if  obj.categoria_laboral is not None:
            categoria= obj.categoria_laboral.nome

        if  obj.funcao_chefia is not None:
            funcao_chefia= obj.funcao_chefia.nome

        if  obj.direccao is not None:
            direccao= obj.direccao.nome

        worksheet.append([
        obj.numero_mecanografico,
        obj.numero_seguranca_social,
        obj.firstname,
        obj.gender,
        obj.date_of_birth.strftime("%d-%m-%Y") if obj.date_of_birth else '',
        obj.personnel_number,
        obj.data_de_emissao,
        obj.data_de_validade,
        obj.estado_civil,
        obj.localidade,
        obj.morada,
        obj.provincia_residencia,
        obj.provincia_nascimento,
        obj.naturalidade,
        obj.telefone,
        obj.correio_electronico,
        obj.data_de_admissao.strftime("%d-%m-%Y") if obj.data_de_admissao else '',
        direccao,
        obj.estabelecimento,
        categoria,
        # obj.categoria_nova.nome,
        funcao_chefia,
        obj.vencimento_mensal,
        #obj.numero_dependentes,
        obj.current_status,
        # obj.estado,
        obj.habilitacao,
        obj.area_de_formacao,
        ])
        # obj.estado_pv,
        # obj.data_prova_vida.strftime("%d-%m-%Y"),
        # obj.observacao
        

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_funcionarios_activos.xlsx'
    workbook.save(response)

    return response

def export_to_excel_licenciados(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Número Mecanográfico", 
    "Número INSS", 
    # "Novo Número Mecanográfico", 
    "Nome Completo", 
    "Gênero", 
    "Data de Nascimento", 
    "Nº do BI", 
    "Data de Emissão", 
    "Data de Validade",
    "Estado Cívil", 
    "Localidade", 
    "Morada", 
    "Província de Residência",
    "Província de Nascimento", 
    "Naturalidade", 
    "Contacto Telefônico",
    "Correio Electrônico",  
    "Data de Admissão", 
    "Departamento", 
    "Estabelecimento", 
    "Categoria Laboral", 
    # "Nova Categoria Laboral", 
    "Função de Chefia", 
    #"Número de Dependentes", 
    "Vencimento Mensal", 
    # "Novo Vencimento Mensal", 
    "Situação", 
    # "Estado", 
    "Habilitação Literária",
    "Área de Formação"
    ])  
    # "Estado da Prova de Vida", 
    # "Data de Realização da Prova de Vida", 
    # "Observações"   # Adding headers

    queryset = Employee.objects.filter(current_status='licenciado').order_by("firstname")  # Replace 'YourModel' with the name of your Django model

    for obj in queryset:

        categoria_laboral = ''
        funcao_chefia = ''
        direccao = ''
        
        if  obj.categoria_laboral is not None:
            categoria= obj.categoria_laboral.nome

        if  obj.funcao_chefia is not None:
            funcao_chefia= obj.funcao_chefia.nome

        if  obj.direccao is not None:
            direccao= obj.direccao.nome

        worksheet.append([
        obj.numero_mecanografico,
        obj.numero_seguranca_social,
        obj.firstname,
        obj.gender,obj.date_of_birth.strftime("%d-%m-%Y"),
        obj.personnel_number,
        obj.data_de_emissao,
        obj.data_de_validade,
        obj.estado_civil,
        obj.localidade,
        obj.morada,
        obj.provincia_residencia,
        obj.provincia_nascimento,
        obj.naturalidade,
        obj.telefone,
        obj.correio_electronico,
        obj.data_de_admissao.strftime("%d-%m-%Y"),
        direccao,
        obj.estabelecimento,
        categoria,
        # obj.categoria_nova.nome,
        funcao_chefia,
        obj.vencimento_mensal,
        #obj.numero_dependentes,
        obj.current_status,
        # obj.estado,
        obj.habilitacao,
        obj.area_de_formacao,
        ])
        # obj.estado_pv,
        # obj.data_prova_vida.strftime("%d-%m-%Y"),
        # obj.observacao
        

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_funcionarios_licenciados.xlsx'
    workbook.save(response)

    return response

def export_to_excel_pre_reformados(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Número Mecanográfico", 
    "Número INSS", 
    # "Novo Número Mecanográfico", 
    "Nome Completo", 
    "Gênero", 
    "Data de Nascimento", 
    "Nº do BI", 
    "Data de Emissão", 
    "Data de Validade",
    "Estado Cívil", 
    "Localidade", 
    "Morada", 
    "Província de Residência",
    "Província de Nascimento", 
    "Naturalidade", 
    "Contacto Telefônico",
    "Correio Electrônico",  
    "Data de Admissão", 
    "Departamento", 
    "Estabelecimento", 
    "Categoria Laboral", 
    # "Nova Categoria Laboral", 
    "Função de Chefia", 
    #"Número de Dependentes", 
    "Vencimento Mensal", 
    # "Novo Vencimento Mensal", 
    "Situação", 
    # "Estado", 
    "Habilitação Literária",
    "Área de Formação"
    ])  
    # "Estado da Prova de Vida", 
    # "Data de Realização da Prova de Vida", 
    # "Observações"   # Adding headers

    queryset = Employee.objects.filter(current_status='pre-reformado').order_by("firstname")  # Replace 'YourModel' with the name of your Django model

    for obj in queryset:

        categoria_laboral = ''
        funcao_chefia = ''
        direccao = ''
        
        if  obj.categoria_laboral is not None:
            categoria= obj.categoria_laboral.nome

        if  obj.funcao_chefia is not None:
            funcao_chefia= obj.funcao_chefia.nome

        if  obj.direccao is not None:
            direccao= obj.direccao.nome

        worksheet.append([
        obj.numero_mecanografico,
        obj.numero_seguranca_social,
        obj.firstname,
        obj.gender,obj.date_of_birth.strftime("%d-%m-%Y"),
        obj.personnel_number,
        obj.data_de_emissao,
        obj.data_de_validade,
        obj.estado_civil,
        obj.localidade,
        obj.morada,
        obj.provincia_residencia,
        obj.provincia_nascimento,
        obj.naturalidade,
        obj.telefone,
        obj.correio_electronico,
        obj.data_de_admissao.strftime("%d-%m-%Y"),
        direccao,
        obj.estabelecimento,
        categoria,
        # obj.categoria_nova.nome,
        funcao_chefia,
        obj.vencimento_mensal,
        #obj.numero_dependentes,
        obj.current_status,
        # obj.estado,
        obj.habilitacao,
        obj.area_de_formacao,
        ])
        # obj.estado_pv,
        # obj.data_prova_vida.strftime("%d-%m-%Y"),
        # obj.observacao
        

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_funcionarios_pre_reformados.xlsx'
    workbook.save(response)

    return response






def export_to_excel_pre_faltoso(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Número Mecanográfico", 
    "Número INSS", 
    # "Novo Número Mecanográfico", 
    "Nome Completo", 
    "Gênero", 
    "Data de Nascimento", 
    "Nº do BI", 
    "Data de Emissão", 
    "Data de Validade",
    "Estado Cívil", 
    "Localidade", 
    "Morada", 
    "Província de Residência",
    "Província de Nascimento", 
    "Naturalidade", 
    "Contacto Telefônico",
    "Correio Electrônico",  
    "Data de Admissão", 
    "Departamento", 
    "Estabelecimento", 
    "Categoria Laboral", 
    # "Nova Categoria Laboral", 
    "Função de Chefia", 
    #"Número de Dependentes", 
    "Vencimento Mensal", 
    # "Novo Vencimento Mensal", 
    "Situação", 
    # "Estado", 
    "Habilitação Literária",
    "Área de Formação"
    ])   
    # "Estado da Prova de Vida", 
    # "Data de Realização da Prova de Vida", 
    # "Observações"   # Adding headers

    queryset = Employee.objects.filter(estado_pv='faltoso').order_by("firstname")  # Replace 'YourModel' with the name of your Django model

    for obj in queryset:

        categoria_laboral = ''
        funcao_chefia = ''
        direccao = ''
        
        if  obj.categoria_laboral is not None:
            categoria= obj.categoria_laboral.nome

        if  obj.funcao_chefia is not None:
            funcao_chefia= obj.funcao_chefia.nome

        if  obj.direccao is not None:
            direccao= obj.direccao.nome

        worksheet.append([
        obj.numero_mecanografico,
        obj.numero_seguranca_social,
        obj.firstname,
        obj.gender,obj.date_of_birth.strftime("%d-%m-%Y"),
        obj.personnel_number,
        obj.data_de_emissao,
        obj.data_de_validade,
        obj.estado_civil,
        obj.localidade,
        obj.morada,
        obj.provincia_residencia,
        obj.provincia_nascimento,
        obj.naturalidade,
        obj.telefone,
        obj.correio_electronico,
        obj.data_de_admissao.strftime("%d-%m-%Y"),
        direccao,
        obj.estabelecimento,
        categoria,
        # obj.categoria_nova.nome,
        funcao_chefia,
        obj.vencimento_mensal,
        #obj.numero_dependentes,
        obj.current_status,
        # obj.estado,
        obj.habilitacao,
        obj.area_de_formacao,
        ])
        # obj.estado_pv,
        # obj.data_prova_vida.strftime("%d-%m-%Y"),
        # obj.observacao
        

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_funcionarios_pre_reformados.xlsx'
    workbook.save(response)

    return response