import csv
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.forms import widgets
from openpyxl import load_workbook
from datetime import datetime
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.views.generic import DetailView, ListView, View
from django.shortcuts import HttpResponseRedirect, redirect, render
from django.views.generic.edit import CreateView, DeleteView, UpdateView
from datetime import datetime
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter, landscape
# from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from django.http import HttpResponse
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageTemplate, Frame, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import landscape, letter, A3
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image, Spacer
from openpyxl import Workbook
from openpyxl import load_workbook
from django.shortcuts import render
from .forms import ExcelUploadForm
from .models import Employee, ProvaVida, Abertura_Prova_Vida, CategoriaNova
from django.core.exceptions import ObjectDoesNotExist
from .models import Abertura_Prova_Vida

from apps.corecode.models import Categoria, DirecaoAlocacao, CategoriaNova, FuncaoChefia

#from apps.finance.models import Invoice

from django.contrib.auth.models import User

from .models import ProvaVida,Employee,Abertura_Prova_Vida

from .forms import (
    ExcelUploadForm,
    PrivaVidaForm
)



class ProvaVidaListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = ProvaVida
    # context_object_name = 'ProvaVida_list'   # your own name for the list as a template variable
    template_name = "prova_vida/prova_vida_list.html" # own template name/location

    """ def get_context_data(request):
        context = super().get_context_data(**kwargs)
        context["form"] = ProvaVidaForm()
        context ={}
        context["ProvaVidas"] = ProvaVida.objects.all()
        return context """

    def get_context_data(self, **kwargs):
        context = super(ProvaVidaListView, self).get_context_data(**kwargs)
        # context["ProvaVidas"] = ProvaVida.objects.filter(current_status='active') # Get 5 ProvaVidas with status 'active'
        context["ProvasVida"] = ProvaVida.objects.all()
        return context

@login_required(login_url="/accounts/login/")
def getFunc(request):
 
    return render(request, "prova_vida/prova_vida_get_func.html")

class ProvaVidaDetailView(LoginRequiredMixin, DetailView):
    model = ProvaVida
    template_name = "prova_vida/prova_vida_detail.html"

    def get_context_data(self, **kwargs):
        context = super(ProvaVidaDetailView, self).get_context_data(**kwargs)
    #    context["payments"] = Invoice.objects.filter(ProvaVida=self.object)
        return context


class ProvaVidaCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = ProvaVida
    fields = "__all__"
    success_message = "Новый сотрудник успешно добавлен."

    def get_form(self):
        """add date picker in forms"""
        form = super(ProvaVidaCreateView, self).get_form()
        form.fields["data_prova_vida"].widget = widgets.DateInput(attrs={"type": "date"})
        return form


class ProvaVidaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = ProvaVida
    fields = "__all__"
    success_message = "Prova de vida Actualizada Com Sucesso!"

    def get_form(self):
        """add date picker in forms"""
        form = super(ProvaVidaUpdateView, self).get_form()
        form.fields["data_prova_vida"].widget = widgets.DateInput(attrs={"type": "date"})
  

        return form


class ProvaVidaDeleteView(LoginRequiredMixin, DeleteView):
    model = ProvaVida
    success_url = reverse_lazy("prova-vida-list")



class ProvaGetFuncDetailView(LoginRequiredMixin, DetailView):
    model = Employee
    template_name = "prova_vida/prova_vida_detail_func.html"

    def get_context_data(self, **kwargs):
        context = super(ProvaGetFuncDetailView, self).get_context_data(**kwargs)
    #    context["payments"] = Invoice.objects.filter(ProvaVida=self.object)
        return context


@login_required(login_url="/accounts/login/")
def provaGetFuncDetail(request) :
    bi=request.GET['bi']

    direccoes = DirecaoAlocacao.objects.all()
    categorias = Categoria.objects.all()
    categorias_novas = CategoriaNova.objects.all()
    funcoes_chefia = FuncaoChefia.objects.all()
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
    
    employee = Employee.objects.filter(personnel_number=bi)
 
    if len(employee)==0:
        employee = Employee.objects.filter(numero_mecanografico=bi)
    
    if len(employee)==0:
        context = {"error": "Não existe nenhum funcionário com este número do BI ou Número Mecanográfico."}
        template_name = "prova_vida/prova_vida_get_func.html"
        return render(request,template_name,context)
    

    

    if len(abertura)>0:
        prova_vida = ProvaVida.objects.filter(funcionario=employee[0],abertura_prova_vida=abertura[0])
        if len(prova_vida) >0:
            context = {"error": "Este funcinário, já fez a prova de vida."}
            template_name = "prova_vida/prova_vida_get_func.html"
            return render(request,template_name,context)
        
    employee[0].data_de_admissao=employee[0].data_de_admissao.strftime('%Y-%m-%d')
    employee[0].date_of_birth=employee[0].date_of_birth.strftime('%Y-%m-%d')
    template_name = "prova_vida/prova_vida_detail_func.html"
    context = {"employee": employee[0],"direccoes":direccoes,"categorias":categorias,"categorias_novas":categorias_novas,"funcoes_chefia":funcoes_chefia,}
    return render(request, template_name, context)
    



@login_required(login_url="/accounts/login/")
def efectuarProvaVida(request) :
    bi=request.POST['bi']

    
    employee = Employee.objects.filter(personnel_number=bi)
    if len(employee)==0:
        employee = Employee.objects.filter(numero_mecanografico=bi)

    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
    

    if len(abertura)==0:
        context = {"error": "Não há nenhuma prova de vida aberta neste momento."}
        template_name = "prova_vida/prova_vida_get_func.html"
        return render(request,template_name,{"error":context})

    elif len(employee)==0:
        context = {"error": "Não existe nenhum funcionário com este número do BI ou número mecanográfico."}
        template_name = "prova_vida/prova_vida_get_func.html"
        return render(request,template_name,context)
    elif len(abertura)>0:
        prova_vida = ProvaVida.objects.filter(funcionario=employee[0],abertura_prova_vida=abertura[0])
        if len(prova_vida) >0:
            context = {"error": "Este funcionário, já fez a prova de vida."}
            template_name = "prova_vida/prova_vida_get_func.html"
            return render(request,template_name,context)
    try:
        categoria_nova = CategoriaNova.objects.get(pk=request.POST['categoria_nova'])
    except:
        categoria_nova=None

    prova_vida=ProvaVida(abertura_prova_vida = abertura[0],funcionario=employee[0],
        novo_numero_mecanografico=request.POST['novo_numero_mecanografico'],
        novo_vencimento=request.POST['outro_vencimento_mensal'],
        categoria_nova = categoria_nova,
        user= request.user,
        observacao = request.POST['observacao']
    
        )


    prova_vida.save()
    employee[0].estado_pv="feito"

    try:
        direcao = DirecaoAlocacao.objects.get(pk=request.POST['direcao_alocacao'])
    except:
        direcao=None

    try:
        categoria = Categoria.objects.get(pk=request.POST['categoria'])
    except:
        categoria=None

    try:
        categoria_nova = CategoriaNova.objects.get(pk=request.POST['categoria_nova'])
    except:
        categoria_nova=None

    try:
        funcao_chefia = FuncaoChefia.objects.get(pk=request.POST['funcao_chefia'])
    except:
        funcao_chefia=None
    


    employee[0].firstname=request.POST['primeiro_nome'] 
    # employee[0].surname=request.POST['sobre_nome']
    #funcionario.current_status = request.POST['reforma'],
    employee[0].personnel_number = request.POST['numero_bi']
    employee[0].numero_mecanografico = request.POST['numero_mecanografico']
    employee[0].gender = request.POST['genero']
    employee[0].date_of_birth = request.POST['data_nascimento']
    #datetime.strptime (request.POST['data_nascimento'],'%Y-%m-%d') ,
    employee[0].provincia_residencia = request.POST['provincia_residencia']
    employee[0].estado_civil = request.POST['estado_civil']
    employee[0].localidade= request.POST['localidade']
    employee[0].telefone =request.POST['telefone']
    employee[0].numero_dependentes =request.POST['numero_dependentes']
    employee[0].morada =request.POST['morada']
    #funcionario.numero_dependentes =request.POST['numero_dependentes']
    if categoria is not None:
        employee[0].categoria_laboral =categoria

    if categoria_nova is not None:
        employee[0].categoria_laboral_nova = categoria_nova

    if funcao_chefia is not None:
        employee[0].funcao_chefia = funcao_chefia

    employee[0].data_de_admissao = request.POST['data_admissao']

    employee[0].estabelecimento = request.POST['estabelecimento']
     
    

    #direccao_alocacao = direcao,

    #citizenship = 'angolana'

    employee[0].vencimento_mensal = request.POST['vencimento_mensal']

    employee[0].current_status = request.POST['reforma']

    # categoria_laboral_antiga = request.POST['categoria_antiga'],
    # outra_categoria_laboral= request.POST['outra_categoria'],

    # funcao_chefia_antiga = request.POST['funcao_chefia'],
    # outra_funcao_chefia= request.POST['outra_funcao'],

    # reforma= request.POST['reforma'],
    employee[0].habilitacao =request.POST['habilitacao']

    # funcionario.numero_telefone = request.POST['telefone'],

    employee[0].save()

    employee[0].data_de_admissao=datetime.strptime (employee[0].data_de_admissao,'%Y-%m-%d')
    employee[0].date_of_birth=datetime.strptime (employee[0].date_of_birth,'%Y-%m-%d')
    

    context = {"prova_vida": prova_vida,"employee":employee[0]}

    template_name = "prova_vida/prova_vida_sucess.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def editarProvaVida(request):

    bi = request.POST['bi']
    
    employee = Employee.objects.filter(personnel_number=bi)
    if len(employee) == 0:
        employee = Employee.objects.filter(numero_mecanografico=bi)

    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")

    if len(abertura) == 0:
        context = {"error": "Não há nenhuma prova de vida aberta neste momento."}
        template_name = "prova_vida/prova_vida_get_func.html"
        return render(request, template_name, {"error": context})

    if len(employee) == 0:
        context = {"error": "Não existe nenhum funcionário com este número do BI ou número mecanográfico."}
        template_name = "prova_vida/prova_vida_get_func.html"
        return render(request, template_name, context)

    prova_vida = ProvaVida.objects.filter(funcionario=employee[0], abertura_prova_vida=abertura[0]).first()

    if prova_vida:
        try:
            direcao = DirecaoAlocacao.objects.get(pk=request.POST['direcao_alocacao'])
        except:
            direcao = None

        try:
            categoria = Categoria.objects.get(pk=request.POST['categoria'])
        except:
            categoria = None

        try:
            funcao_chefia = FuncaoChefia.objects.get(pk=request.POST['funcao_chefia'])
        except:
            funcao_chefia = None

        try:
            categoria_nova = CategoriaNova.objects.get(pk=request.POST['categoria_nova'])
        except:
            categoria_nova = None

       
        # Atualize os campos relevantes do objeto ProvaVida

        prova_vida.novo_numero_mecanografico = request.POST['novo_numero_mecanografico']
        prova_vida.novo_vencimento = request.POST['outro_vencimento_mensal']
        prova_vida.observacao = request.POST['observacao']
        prova_vida.categoria_nova = categoria_nova

        # Salve o objeto ProvaVida atualizado
        prova_vida.save()

        # Atualize os campos relevantes do objeto Employee
        if categoria is not None:
            employee[0].categoria_laboral =categoria


        if funcao_chefia is not None:
            employee[0].funcao_chefia = funcao_chefia

        if direcao is not None:
            employee[0].direccao = direcao

        employee[0].firstname = request.POST['primeiro_nome'] 
        employee[0].personnel_number = request.POST['numero_bi']
        employee[0].numero_mecanografico = request.POST['numero_mecanografico']
        employee[0].gender = request.POST['genero']
        employee[0].date_of_birth = request.POST['data_nascimento']
        employee[0].provincia_residencia = request.POST['provincia_residencia']
        employee[0].estado_civil = request.POST['estado_civil']
        employee[0].localidade = request.POST['localidade']
        employee[0].telefone = request.POST['telefone']
        employee[0].numero_dependentes = request.POST['numero_dependentes']
        employee[0].morada = request.POST['morada']
        employee[0].data_de_admissao = request.POST['data_admissao']
        employee[0].estabelecimento = request.POST['estabelecimento']
        employee[0].vencimento_mensal = request.POST['vencimento_mensal']
        employee[0].current_status = request.POST['reforma']
        employee[0].habilitacao = request.POST['habilitacao']
        employee[0].estado_pv = "feito"

        # Salve o objeto Employee atualizado
        employee[0].save()

        context = {"prova_vida": prova_vida, "employee": employee[0]}
        template_name = "prova_vida/prova_vida_sucess_edit.html"

        return render(request, template_name, context)
    else:
        context = {"error": "Não existe prova de vida para este funcionário e abertura de prova de vida."}
        template_name = "prova_vida/prova_vida_get_func.html"
        return render(request, template_name, context)


@login_required(login_url="/accounts/login/")
def prova_vida_edit(request,id):

    prova_vida = ProvaVida.objects.get(pk=id)

    categorias  = Categoria.objects.filter(estado_objecto='activo')
    categorias_novas  = CategoriaNova.objects.filter(estado_objecto='activo')
    direccoes  = DirecaoAlocacao.objects.filter(estado_objecto='activo')
    funcoes_chefia  = FuncaoChefia.objects.filter(estado_objecto='activo')
   
    
    funcionario = Employee.objects.get(pk=id)
    funcionario.data_de_admissao=funcionario.data_de_admissao.strftime('%Y-%m-%d')
    funcionario.date_of_birth=funcionario.date_of_birth.strftime('%Y-%m-%d')

    context = {"funcionario": funcionario,
            "prova_vida":prova_vida,
            'categorias':categorias,
            'categorias_novas':categorias_novas,
            'direccoes':direccoes,
            'funcoes_chefia':funcoes_chefia,}
    template_name = "prova_vida/prova_vida_edit.html"

    return render(request, template_name, context)




def editFuncionario(request):
    funcionarios  = Employee.objects.filter(estado_objecto='activo')
    categorias  = Categoria.objects.filter(estado_objecto='activo')
    categorias_novas  = CategoriaNova.objects.filter(estado_objecto='activo')
    direcoes  = DirecaoAlocacao.objects.filter(estado_objecto='activo')
    funcoes_chefia  = FuncaoChefia.objects.filter(estado_objecto='activo')
    try:
        direccao = DirecaoAlocacao.objects.get(pk=request.POST['direcao_alocacao'])
    except:
        direccao=None

    try:
        categoria = Categoria.objects.get(pk=request.POST['categoria'])
    except:
        categoria=None

    try:
        categoria_nova = CategoriaNova.objects.get(pk=request.POST['categoria_nova'])
    except:
        categoria_nova=None

    try:
        funcao_chefia = FuncaoChefia.objects.get(pk=request.POST['funcao_chefia'])
    except:
        funcao_chefia=None
    

   

    funcionario = Employee.objects.get(pk=request.POST['id'])
    funcionario.firstname=request.POST['primeiro_nome'] 
    # funcionario.surname=request.POST['sobre_nome']
    #funcionario.current_status = request.POST['reforma'],
    funcionario.personnel_number = request.POST['numero_bi']
    funcionario.numero_mecanografico = request.POST['numero_mecanografico']
    funcionario.gender = request.POST['genero']
    funcionario.date_of_birth = datetime.strptime (request.POST['data_nascimento'],'%Y-%m-%d')
    #datetime.strptime (request.POST['data_nascimento'],'%Y-%m-%d') ,
    funcionario.provincia_residencia = request.POST['provincia_residencia']
    funcionario.estado_civil = request.POST['estado_civil']
    funcionario.localidade= request.POST['localidade']
    funcionario.numero_dependentes= request.POST['numero_dependentes']
    funcionario.morada= request.POST['morada']
    funcionario.telefone =request.POST['telefone']
    #funcionario.numero_dependentes =request.POST['numero_dependentes']
    if categoria is not None:
        funcionario.categoria_laboral =categoria

    if funcao_chefia is not None:
        funcionario.funcao_chefia =funcao_chefia

    if direccao is not None:
        funcionario.direccao =direccao

    """if categoria_nova is not None:
        funcionario.categoria_laboral_nova =categoria_nova"""


    funcionario.data_de_admissao = datetime.strptime (request.POST['data_admissao'],'%Y-%m-%d')

    funcionario.estabelecimento = request.POST['estabelecimento']
    #datetime.strptime (request.POST['data_admissao'],'%Y-%m-%d'),
    

    #direccao_alocacao = direcao,

    #citizenship = 'angolana'

    funcionario.vencimento_mensal = request.POST['vencimento_mensal']

    funcionario.current_status = request.POST['reforma']

    # categoria_laboral_antiga = request.POST['categoria_antiga'],
    # outra_categoria_laboral= request.POST['outra_categoria'],

    # funcionario.funcao_chefia = request.POST['funcao_chefia'],
    # outra_funcao_chefia= request.POST['outra_funcao'],

    # reforma= request.POST['reforma'],
    funcionario.habilitacao =request.POST['habilitacao']

    # funcionario.numero_telefone = request.POST['telefone'],

    funcionario.estado=request.POST['estado']
    funcionario.origem = 'incluso'




    

    funcionario.save()
    

    context = {"funcionario":funcionario}

    template_name = "corecode/edit_success.html"

    return render(request, template_name,context)



@login_required(login_url="/accounts/login/")
def getProvaVidasFunc(request) :
    prova_vida =[]
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
    
    if len(abertura)!=0:
        prova_vida  = ProvaVida.objects.select_related('funcionario')
    
   
    context = {"provas_vida": prova_vida}

    template_name = "prova_vida/prova_vida_list.html"

    return render(request, template_name,context)




@login_required(login_url="/accounts/login/")
def getProvaVidasRelatorio(request) :

    funcionario_concluidos =[]
    funcionario_faltas =[]
    

    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
    
    if len(abertura)!=0:
        funcionario_concluidos=Employee.objects.filter(provavida__abertura_prova_vida=abertura[0]).distinct()

        funcionario_faltas = Employee.objects.exclude(provavida__abertura_prova_vida=abertura[0]).distinct()
    else:
        abertura = Abertura_Prova_Vida.objects.filter(estado_actual="fechada")
        if len(abertura)!=0:
            funcionario_concluidos=Employee.objects.filter(provavida__abertura_prova_vida=abertura[0]).distinct()

            funcionario_faltas = Employee.objects.exclude(provavida__abertura_prova_vida=abertura[0]).distinct()
    

    context = {"funcionario_concluidos": funcionario_concluidos,
               "funcionario_faltas": funcionario_faltas,
               "qtd_conc":len(funcionario_concluidos),
               "qtd_falta":len(funcionario_faltas),
               }

    template_name = "prova_vida/prova_vida_report.html"

    return render(request, template_name,context)





@login_required(login_url="/accounts/login/")
def prova_vida_func_detail(request, id):
    prova_vida = ProvaVida.objects.get(pk=id)

  
    context = {"prova_vida":prova_vida}
    template_name = "prova_vida/prova_vida_func_detail.html"
    return render(request, template_name, context)



class Abertura_Prova_VidaListView(LoginRequiredMixin, SuccessMessageMixin, ListView):
    model = Abertura_Prova_Vida
    # context_object_name = 'Abertura_Prova_Vida_list'   # your own name for the list as a template variable
    template_name = "abertura_prova_vida/abertura_prova_vida_list.html" # own template name/location

    """ def get_context_data(request):
        context = super().get_context_data(**kwargs)
        context["form"] = Abertura_Prova_VidaForm()
        context ={}
        context["Aberturas_Prova_Vida"] = Abertura_Prova_Vida.objects.all()
        return context """

    def get_context_data(self, **kwargs):
        context = super(Abertura_Prova_VidaListView, self).get_context_data(**kwargs)
        # context["Abertura_Prova_Vidas"] = Abertura_Prova_Vida.objects.filter(current_status='active') # Get 5 Abertura_Prova_Vidas with status 'active'
        context["aberturas"] = Abertura_Prova_Vida.objects.all()
        return context


class Abertura_Prova_VidaDetailView(LoginRequiredMixin, DetailView):
    model = Abertura_Prova_Vida
    template_name = "abertura_prova_vida/abertura_prova_vida_detail.html"

    def get_context_data(self, **kwargs):
        context = super(Abertura_Prova_VidaDetailView, self).get_context_data(**kwargs)
    #    context["payments"] = Invoice.objects.filter(Abertura_Prova_Vida=self.object)
        return context


class Abertura_Prova_VidaCreateView(LoginRequiredMixin, SuccessMessageMixin, CreateView):
    model = Abertura_Prova_Vida
    fields = "__all__"
    success_message = "Prova de vida Aberta Com Sucesso."

    def get_form(self):
        """add date picker in forms"""
        form = super(Abertura_Prova_VidaCreateView, self).get_form()
        form.fields["data_de_abertura"].widget = widgets.DateInput(attrs={"type": "date"})
        form.fields["data_de_fim"].widget = widgets.DateInput(
            attrs={"type": "date"}
        )
        return form


class Abertura_Prova_VidaUpdateView(LoginRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Abertura_Prova_Vida
    fields = "__all__"
    success_message = "Actualizado com Sucesso."

    def get_form(self):
        """add date picker in forms"""
        form = super(Abertura_Prova_VidaUpdateView, self).get_form()
        form.fields["data_de_abertura"].widget = widgets.DateInput(attrs={"type": "date"})
        form.fields["date_de_fim"].widget = widgets.DateInput(
            attrs={"type": "date"}
        )
      
        form.fields['photo'].widget = widgets.FileInput()
        return form


class Abertura_Prova_VidaDeleteView(LoginRequiredMixin, DeleteView):
    model = Abertura_Prova_Vida
    success_url = reverse_lazy("abertura_prova_vida-list")



class DownloadCSVViewdownloadcsv(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = 'attachment; filename="Abertura_Prova_Vida_template.csv"'

        writer = csv.writer(response)
        writer.writerow(
            [
                "registration_number",
                "surname",
                "firstname",
                "other_names",
                "gender",
                "parent_number",
                "address",
                "current_class",
            ]
        )

        return response



@login_required(login_url="/accounts/login/")
def fecharProvaVida(request) :
    #bi=request.POST['bi']
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
    if len(abertura)>0:   
        abertura[0].estado_actual="fechada"

   

        abertura[0].save()
        Employee.objects.filter(estado_pv="pendente").update(estado_pv="faltoso")

    context = {"abertura": abertura}

    template_name = "abertura_prova_vida/abertura_prova_vida_fecho_resumo.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def abrirProvaVida(request) :
   # bi=request.POST['bi']
    referencia=request.POST['referencia']
    descricao=request.POST['descricao']
    data_inicio=datetime.strptime (request.POST['datainicio'],'%Y-%m-%d')
    
    data_fim=datetime.strptime (request.POST['datafim'],'%Y-%m-%d')
    
  
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual='aberta')

    if len(abertura)>0:
        context = {"error": "já existe um processo de prova de vida aberta, deve fechar primeiro caso queira abrir outra"}
        template_name = "abertura_prova_vida/abertura_prova_vida_form.html"
        return render(request,template_name,context)
    elif data_inicio.date() > data_fim.date():
        context = {"error": "a Data de fim deve sempre ser maior que a data de inicio"}
        template_name = "abertura_prova_vida/abertura_prova_vida_form.html"


        return render(request,template_name,context)



    abertura_prova_vida=Abertura_Prova_Vida(data_de_abertura=data_inicio,data_de_fim=data_fim,descricao=descricao,referencia=referencia)
    

    abertura_prova_vida.save()

    context = {"abertura": abertura_prova_vida}
    

    template_name = "abertura_prova_vida/abertura_prova_vida_resume.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def provaVidaForm(request) :

    

    template_name = "abertura_prova_vida/abertura_prova_vida_form.html"
    ano=datetime.now().year
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual='fechada')

    referencia = 'PV-'+str(ano)+'-'+str(len(abertura)+1)
    context = {"referencia": referencia}

    return render(request, template_name,context)

@login_required(login_url="/accounts/login/")
def provaVidaReabrirForm (request):
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual='aberta').first()

    if abertura is not  None:
        template_name = "index.html"
        context = {'error':'Há ainda uma prova de vida em curso.'}

        return render(request, template_name,context)

    #context = {"employee": employee}

    abertura = Abertura_Prova_Vida.objects.filter(estado_actual='fechada').order_by('data_de_abertura').last()




    abertura.data_de_abertura = abertura.data_de_abertura.strftime('%Y-%m-%d')
    abertura.data_de_fim = abertura.data_de_fim.strftime('%Y-%m-%d')

    context = {'abertura':abertura}


    template_name = "abertura_prova_vida/reabertura_prova_vida_form.html"

    return render(request, template_name,context)
 

@login_required(login_url="/accounts/login/")
def getProvaVida(request) :
    #bi=request.POST['bi']
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")

    if len(abertura)>0:


        context = {"abertura": abertura[0]}

        template_name = "abertura_prova_vida/abertura_prova_vida_fecho.html"

        return render(request, template_name,context)
    context = {"error": "não existe um processo de prova de vida aberta"}
    template_name = "index.html"
    return render(request,template_name,context)


@login_required(login_url="/accounts/login/")
def getProvaVidaProrrogar(request) :
 
    abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")

    if len(abertura)>0:


        context = {"abertura": abertura[0]}

        template_name = "abertura_prova_vida/abertura_prova_vida_fecho.html"

        return render(request, template_name,context)
    context = {"error": "não existe um processo de prova de vida aberta"}
    template_name = "index.html"
    return render(request,template_name,context)

@login_required(login_url="/accounts/login/")
def getProvaVidaReabrir(request) :

    ultima_abertura= Abertura_Prova_Vida.objects.filter(estado_actual="fechada").order_by('data_de_abertura').last()

    if ultima_abertura is not None:


        context = {"ultima_abertura": ultima_abertura}

        template_name = "abertura_prova_vida/abertura_prova_vida_fecho.html"

        return render(request, template_name,context)
    context = {"error": "não existe um processo de prova de vida aberta"}
    template_name = "index.html"
    return render(request,template_name,context)


@login_required(login_url="/accounts/login/")
def reabrirProvaVida(request) :
    id=request.POST['id']
    abertura = Abertura_Prova_Vida.objects.get(pk=id)
    abertura.data_de_fim= datetime.strptime (request.POST['datafim'],'%Y-%m-%d')
    abertura.estado_actual='aberta'

    abertura.save()
    Employee.objects.filter(estado_pv="faltoso").update(estado_pv="pendente")
    context = {"abertura": abertura}

    template_name = "abertura_prova_vida/abertura_prova_vida_reabertura_resumo.html"

    return render(request, template_name,context)


@login_required(login_url="/accounts/login/")
def export_to_excel(request):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Número Mecanográfico", 
    "Novo Número Mecanográfico", 
    "Nome Completo", "Gênero", 
    "Data de Nascimento", 
    "Nº do BI", 
    "Estado Cívil", 
    "Localidade", 
    "Morada", 
    "Província de Residência", 
    "Contacto Telefônico", 
    "Data de Admissão", 
    "Direcção de Alocação", 
    "Estabelecimento", 
    "Categoria Laboral", 
    "Nova Categoria Laboral", 
    "Função de Chefia", 
    "Número de Dependentes", 
    "Vencimento Mensal", 
    "Novo Vencimento Mensal", 
    "Situação", 
    "Estado", 
    "Habilitação Literária", 
    "Estado da Prova de Vida", 
    "Data de Realização da Prova de Vida", 
    "Observações" ])  # Adding headers

    queryset = ProvaVida.objects.all()  # Replace 'YourModel' with the name of your Django model

    for obj in queryset:

        categoria_laboral = ''
        categoria_nova = ''
        funcao_chefia = ''
        direccao = ''

        if  obj.funcionario.categoria_laboral is not None:
            categoria= obj.funcionario.categoria_laboral.nome

        if  obj.funcionario.funcao_chefia is not None:
            funcao_chefia= obj.funcionario.funcao_chefia.nome

        if  obj.categoria_nova is not None:
            categoria_nova = obj.categoria_nova.nome 


        if  obj.funcionario.direccao is not None:
            direccao= obj.funcionario.direccao.nome


        worksheet.append([obj.funcionario.numero_mecanografico,obj.novo_numero_mecanografico,
        obj.funcionario.firstname,
        obj.funcionario.gender,obj.funcionario.date_of_birth.strftime("%d-%m-%Y"),
        obj.funcionario.personnel_number,
        obj.funcionario.estado_civil,
        obj.funcionario.localidade,
        obj.funcionario.morada,
        obj.funcionario.provincia_residencia,obj.funcionario.telefone,
        obj.funcionario.data_de_admissao.strftime("%d-%m-%Y"),
        direccao,
        obj.funcionario.estabelecimento,
        categoria,
        categoria_nova,
        funcao_chefia,
        obj.funcionario.numero_dependentes,
        obj.funcionario.vencimento_mensal,
        obj.novo_vencimento,
        obj.funcionario.current_status,obj.funcionario.estado,
        obj.funcionario.habilitacao,
        obj.funcionario.estado_pv,
        obj.data_prova_vida.strftime("%d-%m-%Y"),
        obj.observacao])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_provas_de_vida_concluidas.xlsx'
    workbook.save(response)

    return response


def export_to_excel_faltoso(request):

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Número Mecanográfico", 
    # "Novo Número Mecanográfico", 
    "Nome Completo", "Gênero", 
    "Data de Nascimento", 
    "Nº do BI", 
    "Estado Cívil", 
    "Localidade", 
    "Morada", 
    "Província de Residência", 
    "Contacto Telefônico", 
    "Data de Admissão", 
    "Direcção de Alocação", 
    "Estabelecimento", 
    "Categoria Laboral", 
    # "Nova Categoria Laboral", 
    "Função de Chefia", 
    #"Número de Dependentes", 
    "Vencimento Mensal", 
    # "Novo Vencimento Mensal", 
    "Situação", 
    "Estado", 
    "Habilitação Literária", 
    "Estado da Prova de Vida"]) 
    # "Data de Realização da Prova de Vida", 
    # "Observações"  # Adding headers

    queryset = Employee.objects.filter(estado_pv="faltoso").order_by("firstname")  # Replace 'YourModel' with the name of your Django model

    for obj in queryset:

        categoria_laboral = ''
        funcao_chefia = ''
        direccao = ''
        
        if  obj.categoria_laboral is not None:
            categoria= obj.categoria_laboral.nome

        if  obj.funcao_chefia is not None:
            funcao_chefia= obj.funcao_chefia.nome

        if  obj.direccao is not None:
            direccao= obj.direccao.nome


        worksheet.append([obj.numero_mecanografico,
        # obj.novo_numero_mecanografico,
        obj.firstname,
        obj.gender,
        obj.date_of_birth.strftime("%d-%m-%Y"),
        obj.personnel_number,
        obj.estado_civil,
        obj.localidade,
        obj.morada,
        obj.provincia_residencia,obj.telefone,
        obj.data_de_admissao.strftime("%d-%m-%Y"),
        direccao,
        obj.estabelecimento,
        categoria_laboral,
        # obj.categoria_nova,
        funcao_chefia,
        obj.vencimento_mensal,
        #obj.numero_dependentes,
        obj.current_status,
        obj.estado,
        obj.habilitacao,
        obj.estado_pv])
        # obj.data_prova_vida.strftime("%d-%m-%Y"),
        # obj.observacao

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lista_de_provas_de_vida_em_falta.xlsx'
    workbook.save(response)

    return response



@login_required(login_url="/accounts/login/")
def gerarProvaVidaPDF(request):

    data1 = [
        ['Bold Text', 'Italic Text', 'Normal Text'],
        ['Bold and Italic Text', 'Bold and Italic Text', 'Italic Text'],
        ['Normal Text', 'Bold Text', 'Italic Text']
    ]

    # Define styles
    styles = getSampleStyleSheet()
    bold_style = styles["BodyText"]
    bold_style.fontName = "Helvetica-Bold"
    italic_style = styles["BodyText"]
    italic_style.fontName = "Helvetica-Oblique"
    # Fetch employee data from the database
    prova_vidas   = ProvaVida.objects.all().order_by("data_prova_vida")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=prova_vida_concluidas.pdf'

    doc = SimpleDocTemplate(response, pagesize=A3)
    elements = []

    # Add image at top center
    right_image_path = "./static/imagens/core/left.jpg"
    image_path = "./static/imagens/core/right.jpg"  # Replace with the path to your image
    image = Image(image_path, width=70, height=70)
    image1 = Image(right_image_path, width=200, height=100)
    image.hAlign = 'LEFT'
    image1.hAlign = 'RIGHT'
    #doc.drawString(30,750,'OFFICIAL COMMUNIQUE')
    p=Paragraph("LISTA TOTAL DE PROVAS DE VIDA CONCLUÍDAS",bold_style)
    #p.hAlign='RIGHT'
    #elements.append(image
    elements.append(image1)
    elements.append(p)

    elements.append(Spacer(1, 12))  # Add some space between image and table

    # Table data
    #queryset = YourModel.objects.all()  # Replace 'YourModel' with the name of your Django model
    data = [["Nome", "Nº BI", "Nº Mecanográfico", "Género", "Data de Nascimento", "Estado"]]  # Table headers

    for prova_vida in prova_vidas:
        data.append([prova_vida.funcionario.firstname,
            prova_vida.funcionario.personnel_number,
            prova_vida.funcionario.numero_mecanografico,
            prova_vida.funcionario.gender,
            str(prova_vida.funcionario.date_of_birth),
            prova_vida.funcionario.estado_pv])

    # Create table

    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(table)

    doc.build(elements)

    return response


@login_required(login_url="/accounts/login/")
def gerarProvaVidaFaltosoPDF(request):

    data1 = [
        ['Bold Text', 'Italic Text', 'Normal Text'],
        ['Bold and Italic Text', 'Bold and Italic Text', 'Italic Text'],
        ['Normal Text', 'Bold Text', 'Italic Text']
    ]

    # Define styles
    styles = getSampleStyleSheet()
    bold_style = styles["BodyText"]
    bold_style.fontName = "Helvetica-Bold"
    italic_style = styles["BodyText"]
    italic_style.fontName = "Helvetica-Oblique"
    # Fetch employee data from the database
    funcionarios   = Employee.objects.filter(estado_pv="faltoso").order_by("firstname")

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=prova_vida_faltosos.pdf'

    doc = SimpleDocTemplate(response, pagesize=A3)
    elements = []

    # Add image at top center
    right_image_path = "./static/imagens/core/left.jpg"
    image_path = "./static/imagens/core/right.jpg"  # Replace with the path to your image
    image = Image(image_path, width=70, height=70)
    image1 = Image(right_image_path, width=200, height=100)
    image.hAlign = 'LEFT'
    image1.hAlign = 'RIGHT'
    #doc.drawString(30,750,'OFFICIAL COMMUNIQUE')
    p=Paragraph("LISTA TOTAL DE PROVAS DE VIDA EM FALTA",bold_style)
    #p.hAlign='RIGHT'
    #elements.append(image
    elements.append(image1)
    elements.append(p)

    elements.append(Spacer(1, 12))  # Add some space between image and table

    # Table data
    #queryset = YourModel.objects.all()  # Replace 'YourModel' with the name of your Django model
    data = [["Nome", "Nº BI", "Nº Mecanográfico", "Género", "Data de Nascimento", "Estado"]]  # Table headers

    for funcionario in funcionarios:
        data.append([funcionario.firstname,
            funcionario.personnel_number,
            funcionario.numero_mecanografico,
            funcionario.gender,
            str(funcionario.date_of_birth),
            funcionario.estado_pv])

    # Create table

    table = Table(data)
    table.setStyle(TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
    ('GRID', (0, 0), (-1, -1), 1, colors.black)]))
    elements.append(table)

    doc.build(elements)

    return response


# def upload_excel_prova_vida(request):
#     errors=[]
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES['excel_file']
#             wb = load_workbook(excel_file)
#             ws = wb.active
#             for row in ws.iter_rows(min_row=2, values_only=True):
#                 numero_mecanografico = row[0]
#                 novo_numero_mecanografico=row[1]
#                 nome = row[2]
#                 genero = row[3]
#                 date_nasciemnto = str(datetime.strptime(row[4], "%d-%m-%Y"))
#                 numero_bi = row[5]
#                 estado_civil=  row[6]
#                 localidade =  row[7]
#                 morada =  row[8]
#                 provincia_residencia =  row[9]
#                 estabelecimento =  row[10]
#                 categoria =  row[11]
#                 categoria_nova =  row[12]
#                 funcao_chefia =  row[13]
#                 vencimento_mensal =  row[14]
#                 novo_vencimento =  row[15]
#                 situacao =  row[16]
#                 estado =  row[17]
#                 habilitacao_literaria =  row[18]
#                 estado_pv =  "feito"
#                 data_prova_vida =  row[20]
#                 observacao  =  row[21]
            
#                 employee = Employee.objects.filter(personnel_number=numero_bi)
#                 if len(employee)==0:
#                     employee = Employee.objects.filter(numero_mecanografico=numero_mecanografico)

#                 abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
                

#                 if len(abertura)==0:
#                     context = {"error": "Não há nenhuma prova de vida aberta neste momento."}
#                     template_name = "prova_vida/prova_vida_get_func.html"
#                     return render(request,template_name,{"error":context})

#                 elif len(employee)==0:
#                     #context = {"error": "Não existe nenhum funcionário com este número do BI ou número mecanográfico."}
#                     #template_name = "prova_vida/prova_vida_get_func.html"
#                     #return render(request,template_name,context)
#                     print("Não existe nenhum funcionário com este número do BI ou número mecanográfico")
#                     continue
#                 elif len(abertura)>0:
#                     prova_vida = ProvaVida.objects.filter(funcionario=employee[0],abertura_prova_vida=abertura[0])
#                     if len(prova_vida) >0:
#                         """context = {"error": "Este funcionário, já fez a prova de vida."}
#                         template_name = "prova_vida/prova_vida_get_func.html"
#                         return render(request,template_name,context)"""
#                         #print("Este funcionário, já fez a prova de vida.")
#                         continue
#                 try:
#                     categoria_nova = CategoriaNova.objects.get(pk=categoria_nova)
#                 except:
#                     categoria_nova=None

#                 try:

#                     prova_vida=ProvaVida(abertura_prova_vida = abertura[0],funcionario=employee[0],
#                         novo_numero_mecanografico=novo_numero_mecanografico,
#                         novo_vencimento=novo_vencimento,
#                         data_prova_vida=data_prova_vida,
                    
#                         categoria_nova = categoria_nova,
#                         user= request.user,
#                         observacao = observacao
                    
#                         )


#                     prova_vida.save()
#                     employee[0].estado_pv=estado_pv
#                 except:
#                     print("Prova de Vida Não Inserida!")
#                     continue

#             """try:
#                     direcao = DirecaoAlocacao.objects.get(pk=request.POST['direcao_alocacao'])
#             except:
#                     direcao=None"""

#             try:
#                 categoria = Categoria.objects.get(pk=categoria)
#             except:
#                 categoria=None

#             try:
#                 categoria_nova = CategoriaNova.objects.get(pk=categoria_nova)
#             except:
#                 categoria_nova=None

#             try:
#                 funcao_chefia = FuncaoChefia.objects.get(pk=funcao_chefia)
#             except:
#                 funcao_chefia=None



#             employee[0].firstname=nome
#             employee[0].personnel_number = numero_bi
#             employee[0].numero_mecanografico = numero_mecanografico
#             employee[0].gender = genero
#             employee[0].date_of_birth = datetime.strptime(str(date_nasciemnto), '%Y-%m-%d %H:%M:%S')
#             employee[0].provincia_residencia = provincia_residencia
#             employee[0].estado_civil = estado_civil
#             employee[0].localidade= localidade
            
#             employee[0].morada =morada
#             if categoria is not None:
#                 employee[0].categoria_laboral =categoria

#             if categoria_nova is not None:
#                 employee[0].categoria_nova = categoria_nova

#             if funcao_chefia is not None:
#                 employee[0].funcao_chefia = funcao_chefia

#             employee[0].estabelecimento = estabelecimento


#             employee[0].vencimento_mensal = vencimento_mensal

#             employee[0].current_status = estado
#             employee[0].habilitacao =habilitacao_literaria


#             employee[0].save()


#             return render(request, 'employees/succes.html',{'errors':errors,"msg":"provas de vidas importadas  com sucesso","tipo":"importação de prova de vida"})




# def upload_excel_prova_vida(request):
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES['excel_file']
#             wb = load_workbook(excel_file)
#             ws = wb.active
#             for row in ws.iter_rows(min_row=2, values_only=True):
#                 numero_bi = row[5]
#                 numero_mecanografico = row[0]
#                 novo_numero_mecanografico = row[1]
#                 nome = row[2]
#                 genero = row[3]
#                 date_nascimento = str(datetime.strptime(row[4], "%d-%m-%Y"))
#                 estado_civil = row[6]
#                 localidade = row[7]
#                 morada = row[8]
#                 provincia_residencia = row[9]
#                 telefone = row[10]
#                 data_admissao = row[11]
#                 direcao_alocacao = [12]
#                 estabelecimento = row[13]
#                 categoria = row[14]
#                 categoria_nova = row[15]
#                 funcao_chefia = row[16]
#                 vencimento_mensal = row[17]
#                 novo_vencimento = row[18]
#                 situacao = row[19]
#                 estado = row[20]
#                 habilitacao_literaria = row[21]
#                 estado_pv = row[22]
#                 data_prova_vida = row[23]
#                 observacao = row[24]

#                 # Verificar se o funcionário já existe na base de dados
#                 employee = Employee.objects.filter(personnel_number=numero_bi)
#                 if len(employee) == 0:
#                     employee = Employee.objects.filter(numero_mecanografico=numero_mecanografico)

#                 abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
#                 if len(abertura) == 0:
#                     context = {"error": "Não há nenhuma prova de vida aberta neste momento."}
#                     return render(request, "prova_vida/prova_vida_get_func.html", {"error": context})

#                 elif len(employee) == 0:
#                     print("Não existe nenhum funcionário com este número do BI ou número mecanográfico")
#                     continue

#                 elif len(abertura) > 0:
#                     prova_vida = ProvaVida.objects.filter(funcionario=employee[0], abertura_prova_vida=abertura[0])
#                     if len(prova_vida) > 0:
#                         continue

#                 # try:
#                 #     categoria_nova_obj = CategoriaNova.objects.get(pk=categoria_nova)
#                 # except CategoriaNova.DoesNotExist:
#                 #     categoria_nova_obj = None

#                 try:
#                     prova_vida = ProvaVida.objects.create(
#                         abertura_prova_vida=abertura[0],
#                         funcionario=employee[0],
#                         novo_numero_mecanografico=novo_numero_mecanografico,
#                         novo_vencimento=novo_vencimento,
#                         data_prova_vida=data_prova_vida,
#                         categoria_nova=categoria_nova,
#                         user=request.user,
#                         observacao=observacao
#                     )
#                     employee[0].estado_pv = estado_pv
#                     prova_vida.save()
#                 except Exception as e:
#                     print(f"Erro ao inserir prova de vida: {e}")
#                     continue

#                 try:
#                     categoria_obj = Categoria.objects.get(pk=categoria)
#                 except Categoria.DoesNotExist:
#                     categoria_obj = None

#                 try:
#                     funcao_chefia_obj = FuncaoChefia.objects.get(pk=funcao_chefia)
#                 except FuncaoChefia.DoesNotExist:
#                     funcao_chefia_obj = None

#                 try:
#                     direcao_obj = DirecaoAlocacao.objects.get(pk=direcao_alocacao)
#                 except DirecaoAlocacao.DoesNotExist:
#                     direcao_obj = None

               
                    
#                 employee[0].firstname = nome
#                 employee[0].personnel_number = numero_bi
#                 employee[0].numero_mecanografico = numero_mecanografico
#                 employee[0].gender = genero
#                 employee[0].date_of_birth = date_nascimento
#                 employee[0].provincia_residencia = provincia_residencia
#                 employee[0].estado_civil = estado_civil
#                 employee[0].localidade = localidade
#                 employee[0].morada = morada
#                 employee[0].categoria_laboral = categoria_obj
#                 employee[0].categoria_nova = categoria_nova
#                 employee[0].funcao_chefia = funcao_chefia_obj
#                 employee[0].estabelecimento = estabelecimento
#                 employee[0].vencimento_mensal = vencimento_mensal
#                 employee[0].current_status = estado
#                 employee[0].habilitacao = habilitacao_literaria
#                 employee[0].data_de_admissao = data_admissao
#                 employee[0].telefone = telefone
#                 employee[0].direccao = direcao_obj
#                 employee[0].save()

#             return render(request, 'employees/succes.html',{"msg":"provas de vidas importadas  com sucesso","tipo":"importação de prova de vida"})
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'upload_form.html', {'form': form})


# def upload_excel_prova_vida(request):
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES['excel_file']
#             wb = load_workbook(excel_file)
#             ws = wb.active
#             for row in ws.iter_rows(min_row=2, values_only=True):
#                 numero_bi = row[5]
#                 numero_mecanografico = row[0]
#                 novo_numero_mecanografico = row[1]
#                 nome = row[2]
#                 genero = row[3]
#                 date_nascimento = str(datetime.strptime(row[4], "%d-%m-%Y"))
#                 estado_civil = row[6]
#                 localidade = row[7]
#                 morada = row[8]
#                 provincia_residencia = row[9]
#                 telefone = row[10]
#                 data_admissao = row[11]
#                 direcao_alocacao = [12]
#                 estabelecimento = row[13]
#                 categoria = row[14]
#                 categoria_nova = row[15]
#                 funcao_chefia = row[16]
#                 vencimento_mensal = row[17]
#                 novo_vencimento = row[18]
#                 situacao = row[19]
#                 estado = row[20]
#                 habilitacao_literaria = row[21]
#                 estado_pv = row[22]
#                 data_prova_vida = row[23]
#                 observacao = row[24]

#                 # Verificar se o funcionário já existe na base de dados
#                 employee = Employee.objects.filter(personnel_number=numero_bi)
#                 if len(employee) == 0:
#                     # Adicionar novo funcionário à lista de funcionários
#                     employee = Employee.objects.create(
#                         personnel_number=numero_bi,
#                         numero_mecanografico=numero_mecanografico,
#                         firstname=nome,
#                         gender=genero,
#                         date_of_birth=date_nascimento,
#                         provincia_residencia=provincia_residencia,
#                         estado_civil=estado_civil,
#                         localidade=localidade,
#                         morada=morada,
#                         telefone=telefone,
#                         data_de_admissao=data_admissao,
#                         direccao=direcao_alocacao,
#                         estabelecimento=estabelecimento,
#                         categoria_laboral=categoria,
#                         categoria_nova=categoria_nova,
#                         funcao_chefia=funcao_chefia,
#                         vencimento_mensal=vencimento_mensal,
#                         current_status=estado,
#                         habilitacao=habilitacao_literaria
#                     )
#                 else:
#                     # Funcionário já existe, manter os dados existentes
#                     employee = employee[0]

#                 # Criar nova prova de vida concluída
#                 try:
#                     prova_vida = ProvaVida.objects.create(
#                         abertura_prova_vida=Abertura_Prova_Vida.objects.get(estado_actual="aberta"),
#                         funcionario=employee,
#                         novo_numero_mecanografico=novo_numero_mecanografico,
#                         novo_vencimento=novo_vencimento,
#                         data_prova_vida=data_prova_vida,
#                         categoria_nova=categoria_nova,
#                         user=request.user,
#                         observacao=observacao
#                     )
#                     employee.estado_pv = estado_pv
#                     prova_vida.save()
#                 except Exception as e:
#                     print(f"Erro ao inserir prova de vida: {e}")
#                     continue

#             return render(request, 'employees/succes.html',{"msg":"provas de vidas importadas  com sucesso","tipo":"importação de prova de vida"})
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'upload_form.html', {'form': form})



# def upload_excel_prova_vida(request):
#     if request.method == 'POST':
#         form = ExcelUploadForm(request.POST, request.FILES)
#         if form.is_valid():
#             excel_file = request.FILES['excel_file']
#             wb = load_workbook(excel_file)
#             ws = wb.active
#             for row in ws.iter_rows(min_row=2, values_only=True):
#                 numero_bi = row[5]
#                 numero_mecanografico = row[0]
#                 novo_numero_mecanografico = row[1]
#                 nome = row[2]
#                 genero = row[3]
#                 date_nascimento = str(datetime.strptime(row[4], "%d-%m-%Y").date())
#                 # str(datetime.strptime(row[4], "%d-%m-%Y %H:%M:%S").date())
#                 # str(datetime.strptime(row[4], "%d-%m-%Y"))
#                 estado_civil = row[6]
#                 localidade = row[7]
#                 morada = row[8]
#                 provincia_residencia = row[9]
#                 telefone = row[10]

#                 try:
#                     admissao = str(datetime.strptime(row[11], "%d-%m-%Y").date())
#                 except:
#                     admissao = None

#                 data_admissao =  admissao
#                 direcao_alocacao = row[12]
#                 estabelecimento = row[13]
#                 categoria = row[14]
#                 categoria_nova = row[15]
#                 funcao_chefia = row[16]
#                 vencimento_mensal = row[17]
#                 novo_vencimento = row[18]
#                 situacao = row[19]
#                 estado = row[20]
#                 habilitacao_literaria = row[21]
#                 estado_pv = row[22]
#                 data_prova_vida = str(datetime.strptime(row[23], "%d-%m-%Y").date())
#                 observacao = row[24]

#                 # Verificar se o funcionário já existe na base de dados
#                 employee = Employee.objects.filter(personnel_number=numero_bi)
#                 if len(employee) == 0:
#                     employee = Employee.objects.filter(numero_mecanografico=numero_mecanografico)
#                     if len(employee) == 0:
#                         # Adicionar novo funcionário à lista de funcionários
#                         employee = Employee.objects.create(
#                             personnel_number=numero_bi,
#                             numero_mecanografico=numero_mecanografico,
#                             firstname=nome,
#                             gender=genero,
#                             date_of_birth=date_nascimento,
#                             provincia_residencia=provincia_residencia,
#                             estado_civil=estado_civil,
#                             localidade=localidade,
#                             morada=morada,
#                             telefone=telefone,
#                             data_de_admissao=data_admissao,
#                             direccao=direcao_alocacao,
#                             estabelecimento=estabelecimento,
#                             categoria_laboral=categoria,
#                             categoria_nova=categoria_nova,
#                             funcao_chefia=funcao_chefia,
#                             vencimento_mensal=vencimento_mensal,
#                             current_status=estado,
#                             habilitacao=habilitacao_literaria
#                         )
#                     else:
#                         employee = employee[0]
#                 else:
#                     try:
#                         direcao = DirecaoAlocacao.objects.get(pk=direcao_alocacao)
#                     except:
#                         direcao=None

#                     try:
#                         categoria_ = Categoria.objects.get(pk=categoria)
#                     except:
#                         categoria_=None

#                     try:
#                         categoria_nova = CategoriaNova.objects.get(pk=categoria_nova)
#                     except:
#                         categoria_nova=None

#                     try:
#                         funcao_chefia = FuncaoChefia.objects.get(pk=funcao_chefia)
#                     except:
#                         funcao_chefia=None

                    
#                     employee = employee[0]
#                     employee.numero_mecanografico = numero_mecanografico
#                     employee.firstname = nome
#                     employee.gender = genero 
#                     employee.date_of_birth = date_nascimento
#                     employee.provincia_residencia = provincia_residencia
#                     employee.estado_civil = estado_civil
#                     employee.localidade = localidade
#                     employee.morada = morada
#                     employee.telefone = telefone
#                     employee.data_de_admissao = data_admissao
#                     employee.direccao = direcao
#                     employee.estabelecimento = estabelecimento
#                     employee.categoria_laboral = categoria_
#                     employee.categoria_nova = categoria_nova
#                     employee.funcao_chefia = funcao_chefia
#                     employee.vencimento_mensal = vencimento_mensal
#                     employee.current_status = estado
#                     employee.habilitacao = habilitacao_literaria
#                     employee.save()

#                 abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
#                 if len(abertura) == 0:
#                     context = {"error": "Não há nenhuma prova de vida aberta neste momento."}
#                     return render(request, "prova_vida/prova_vida_get_func.html", {"error": context})

#                 # Crie a instância ProvaVida usando o usuário autenticado
#                 prova_vida = ProvaVida.objects.create(
#                     abertura_prova_vida=abertura[0],
#                     funcionario=employee,
#                     novo_numero_mecanografico=novo_numero_mecanografico,
#                     novo_vencimento=novo_vencimento,
#                     data_prova_vida=data_prova_vida,
#                     categoria_nova=categoria_nova,
#                     # user=request.user,
#                     observacao=observacao
#                 )

#                 employee.estado_pv = estado_pv
#                 employee.save()
#                 prova_vida.save()

#             return render(request, 'employees/succes.html',{"msg":"provas de vidas importadas  com sucesso","tipo":"importação de prova de vida"})
#     else:
#         form = ExcelUploadForm()
#     return render(request, 'upload_form.html', {'form': form})


def upload_excel_prova_vida(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            wb = load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                numero_bi = row[5]
                numero_mecanografico = row[0]
                novo_numero_mecanografico = row[1]
                nome = row[2]
                genero = row[3]
                date_nascimento = str(datetime.strptime(row[4], "%d-%m-%Y").date())
                estado_civil = row[6]
                localidade = row[7]
                morada = row[8]
                provincia_residencia = row[9]
                telefone = row[10]

                try:
                    admissao = str(datetime.strptime(row[11], "%d-%m-%Y").date())
                except:
                    admissao = None

                data_admissao =  admissao
                direcao_alocacao = row[12]
                estabelecimento = row[13]
                categoria = row[14]
                categoria_nova = row[15]
                funcao_chefia = row[16]
                vencimento_mensal = row[17]
                novo_vencimento = row[18]
                situacao = row[19]
                estado = row[20]
                habilitacao_literaria = row[21]
                estado_pv = row[22]
                data_prova_vida = str(datetime.strptime(row[23], "%d-%m-%Y").date())
                observacao = row[24]

                # Verificar se o funcionário já existe na base de dados
                try:
                    employee = Employee.objects.get(personnel_number=numero_bi)
                except ObjectDoesNotExist:
                    try:
                        employee = Employee.objects.get(numero_mecanografico=numero_mecanografico)
                    except ObjectDoesNotExist:
                        # Adicionar novo funcionário à lista de funcionários
                        employee = Employee.objects.create(
                            personnel_number=numero_bi,
                            numero_mecanografico=numero_mecanografico,
                            firstname=nome,
                            gender=genero,
                            date_of_birth=date_nascimento,
                            provincia_residencia=provincia_residencia,
                            estado_civil=estado_civil,
                            localidade=localidade,
                            morada=morada,
                            telefone=telefone,
                            data_de_admissao=data_admissao,
                            direccao=direcao_alocacao,
                            estabelecimento=estabelecimento,
                            categoria_laboral=categoria,
                            categoria_nova=categoria_nova,
                            funcao_chefia=funcao_chefia,
                            vencimento_mensal=vencimento_mensal,
                            current_status=estado,
                            habilitacao=habilitacao_literaria,
                            estado_pv=estado_pv  # Adicionando estado_pv ao criar um novo funcionário
                        )
                    else:
                        # Ignorar este funcionário, pois já existe na base de dados
                        continue
                else:
                    # Ignorar este funcionário, pois já existe na base de dados
                    continue

                # Criar uma nova instância de ProvaVida apenas para funcionários recém-adicionados
                abertura = Abertura_Prova_Vida.objects.filter(estado_actual="aberta")
                if len(abertura) == 0:
                    context = {"error": "Não há nenhuma prova de vida aberta neste momento."}
                    return render(request, "prova_vida/prova_vida_get_func.html", {"error": context})

                prova_vida = ProvaVida.objects.create(
                    abertura_prova_vida=abertura[0],
                    funcionario=employee,
                    novo_numero_mecanografico=novo_numero_mecanografico,
                    novo_vencimento=novo_vencimento,
                    data_prova_vida=data_prova_vida,
                    categoria_nova=categoria_nova,
                    user=request.user,
                    observacao=observacao
                )

                employee.estado_pv = estado_pv
                employee.save()
                prova_vida.save()

            return render(request, 'employees/succes.html',{"msg":"provas de vidas importadas  com sucesso","tipo":"importação de prova de vida"})
    else:
        form = ExcelUploadForm()
    return render(request, 'upload_form.html', {'form': form})



